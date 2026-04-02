from flask_mysqldb import MySQL
from werkzeug.security import generate_password_hash, check_password_hash
import os
from flask import jsonify
import pandas as pd
import MySQLdb.cursors
import MySQLdb
from datetime import datetime, date,timedelta
from flask import Flask, render_template, redirect, url_for, request, session, flash, jsonify
import datetime
from functools import wraps
from dotenv import load_dotenv
from flask import Response
from openpyxl import Workbook
from flask import send_file
import tempfile
from decimal import Decimal
from num2words import num2words
load_dotenv()
app = Flask(__name__)
app.secret_key = "hardware_secret"

# MySQL Config
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = os.getenv("DB_PASSWORD")
app.config['MYSQL_DB'] = 'hardware_billing_software'


mysql = MySQL(app)

@app.route("/")
def home():
    return redirect(url_for("login"))

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

def number_to_words(n):
    return num2words(n, lang='en_IN').title()

@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        username = request.form.get("username").strip()
        new_password = request.form.get("password").strip()

        cur = mysql.connection.cursor()

        # check user exists
        cur.execute("SELECT id FROM users WHERE username=%s", (username,))
        user = cur.fetchone()

        if not user:
            flash("Username not found", "danger")
            return redirect(url_for("forgot_password"))

        # hash new password
        hashed = generate_password_hash(new_password)

        # update password
        cur.execute(
            "UPDATE users SET password=%s WHERE username=%s",
            (hashed, username)
        )
        mysql.connection.commit()
        cur.close()

        flash("Password updated. Please login.", "success")
        return redirect(url_for("login"))

    return render_template("forgot_password.html")

def owner_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != "owner":
            flash("Access Denied", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated_function

@app.context_processor
def utility_processor():

    def is_active(path):
        return request.path.startswith(path)
    return dict(is_active=is_active)

def get_or_create_supplier(cur, supplier_id, supplier_name):

    if supplier_id:
        return supplier_id

    if not supplier_name:
        return None

    cur.execute(
        "SELECT id FROM suppliers WHERE name=%s",
        (supplier_name,)
    )

    row = cur.fetchone()

    if row:
        return row["id"]   # ✅ FIXED

    cur.execute(
        "INSERT INTO suppliers (name, created_at) VALUES (%s, NOW())",
        (supplier_name,)
    )

    return cur.lastrowid

@app.route("/create-owner")
def create_owner():
    cur = mysql.connection.cursor()

    cur.execute("SELECT id FROM users WHERE role='owner'")
    if cur.fetchone():
        return "Owner already exists"

    hashed_password = generate_password_hash("owner123")

    cur.execute("""
        INSERT INTO users (name, username, password, role)
        VALUES (%s,%s,%s,%s)
    """, ("Owner","owner",hashed_password,"owner"))

    mysql.connection.commit()
    return "Owner Created"

@app.route("/users")
@login_required
@owner_required
def users_page():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute("SELECT id,name,username,mobile,role,is_active FROM users ORDER BY role DESC")
    users = cur.fetchall()
    cur.close()

    return render_template(
        "users.html",
        users=users,
        role=session.get("role")   # ✅ ADD THIS
    )

@app.route("/add-user", methods=["POST"])
@login_required
@owner_required
def add_user():
    name = request.form.get("name")
    username = request.form.get("username")
    mobile = request.form.get("mobile")
    role = request.form.get("role")
    password = request.form.get("password")

    hashed = generate_password_hash(password)

    cur = mysql.connection.cursor()
    cur.execute("""
        INSERT INTO users (name,username,password,mobile,role)
        VALUES (%s,%s,%s,%s,%s)
    """,(name,username,hashed,mobile,role))

    mysql.connection.commit()
    cur.close()

    flash("User created","success")
    return redirect(url_for("users_page"))

@app.route("/delete-user/<int:user_id>")
@login_required
@owner_required
def delete_user(user_id):
    cur = mysql.connection.cursor()

    # prevent deleting owner
    cur.execute("SELECT role FROM users WHERE id=%s",(user_id,))
    r = cur.fetchone()
    if r and r[0] == "owner":
        flash("Owner cannot be deleted", "danger")
        return redirect(url_for("users_page"))

    # check financial records
    cur.execute("SELECT COUNT(*) FROM staff_payouts WHERE user_id=%s", (user_id,))
    payouts = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM staff_advances WHERE user_id=%s", (user_id,))
    advances = cur.fetchone()[0]

    if payouts > 0 or advances > 0:
        # deactivate instead
        cur.execute("UPDATE users SET is_active=0 WHERE id=%s", (user_id,))
        mysql.connection.commit()
        cur.close()

        flash("User has financial history. Deactivated instead of deleted.", "warning")
        return redirect(url_for("users_page"))

    # safe delete if no history
    cur.execute("DELETE FROM users WHERE id=%s",(user_id,))
    mysql.connection.commit()
    cur.close()

    flash("User deleted", "info")
    return redirect(url_for("users_page"))


@app.route("/update-user/<int:user_id>", methods=["POST"])
@login_required
@owner_required
def update_user(user_id):

    name = request.form.get("name")
    mobile = request.form.get("mobile")

    cur = mysql.connection.cursor()

    cur.execute("""
        UPDATE users
        SET name=%s, mobile=%s
        WHERE id=%s
    """,(name,mobile,user_id))

    mysql.connection.commit()
    cur.close()

    return jsonify({"success":True})

@app.route("/toggle-user/<int:user_id>")
@login_required
@owner_required
def toggle_user(user_id):
    cur = mysql.connection.cursor()

    cur.execute("UPDATE users SET is_active = NOT is_active WHERE id=%s",(user_id,))
    mysql.connection.commit()
    cur.close()

    return redirect(url_for("users_page"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"].strip()

        cur = mysql.connection.cursor()
        cur.execute("SELECT id, username, password, role, is_active FROM users WHERE username=%s", (username,))
        user = cur.fetchone()
        cur.close()

        if user:
            if not user[4]:
                flash("User blocked", "danger")
                return render_template("login.html")

            if check_password_hash(user[2], password):
                session["user_id"] = user[0]
                session["username"] = user[1]
                session["role"] = user[3]
                return redirect(url_for("dashboard"))

        flash("Invalid Username or Password", "danger")

    return render_template("login.html")

@app.route("/dashboard")
@login_required
def dashboard():
    cur = mysql.connection.cursor()

    # =========================
    # 1️⃣ TOTAL SALE TODAY
    # =========================
    cur.execute("""
    SELECT
        IFNULL(
            (SELECT SUM(total_amount)
             FROM bills
             WHERE DATE(created_at)=CURDATE()
             AND status IN ('DONE','BAKI')
            ),0
        )
        +
        IFNULL(
            (SELECT SUM(amount)
             FROM chillar_entries
             WHERE DATE(created_at)=CURDATE()
             AND type='BIKRI'
            ),0
        )
    """)

    total_sale = cur.fetchone()[0] or 0

    # =========================
    # ADVANCE CASH
    # =========================
    cur.execute("""
    SELECT IFNULL(SUM(advance_amount),0)
    FROM advance_bookings
    WHERE DATE(created_at)=CURDATE()
    AND payment_mode='Cash'
    """)
    advance_cash = cur.fetchone()[0] or 0

    # =========================
    # ADVANCE UPI
    # =========================
    cur.execute("""
    SELECT IFNULL(SUM(advance_amount),0)
    FROM advance_bookings
    WHERE DATE(created_at)=CURDATE()
    AND payment_mode!='Cash'
    """)
    advance_upi = cur.fetchone()[0] or 0
    # =========================
    # 2️⃣ UPI TODAY
    # =========================
    cur.execute("""
    SELECT
        IFNULL(
            (SELECT SUM(paid_amount)
             FROM bills
             WHERE upi_account IN ('Owner','Staff')
             AND DATE(created_at)=CURDATE()
             AND status IN ('DONE','BAKI')
            ),0
        )
        +
        IFNULL(
            (SELECT SUM(amount)
             FROM chillar_entries
             WHERE payment_mode IN ('Owner','Staff')
             AND DATE(created_at)=CURDATE()
             AND type='BIKRI'
            ),0
        )
    """)
    upi_collection = float(cur.fetchone()[0] or 0) + float(advance_upi)

    # =========================
    # 3️⃣ JAWAK (TOTAL EXPENSES TODAY)
    # =========================
    cur.execute("""
    SELECT IFNULL(SUM(amount),0)
    FROM supplier_payments
    WHERE payment_method='UPI'
    AND DATE(payment_date)=CURDATE()
    """)
    upi_supplier_payment = cur.fetchone()[0] or 0

    cur.execute("""
    SELECT IFNULL(SUM(amount),0)
    FROM staff_payouts
    WHERE DATE(created_at)=CURDATE()
    """)
    upi_staff_payment = cur.fetchone()[0] or 0

    cur.execute("""
    SELECT IFNULL(SUM(debit_amount),0)
    FROM hamal_ledger
    WHERE DATE(created_at)=CURDATE()
    """)
    upi_hamal_payment = cur.fetchone()[0] or 0

    # =========================
    # FINAL UPI BALANCE
    # =========================
    upi_balance = (
            float(upi_collection)
            - float(upi_supplier_payment)
            - float(upi_staff_payment)
            - float(upi_hamal_payment)
    )

    # SUPPLIER PAYMENTS
    cur.execute("""
    SELECT IFNULL(SUM(amount),0)
    FROM supplier_payments
    WHERE DATE(payment_date)=CURDATE()
    """)
    supplier_payments = cur.fetchone()[0] or 0

    # STAFF PAYOUTS
    cur.execute("""
    SELECT IFNULL(SUM(amount),0)
    FROM staff_payouts
    WHERE DATE(created_at)=CURDATE()
    """)
    staff_payments = cur.fetchone()[0] or 0

    # HAMAL PAYMENTS
    cur.execute("""
    SELECT IFNULL(SUM(debit_amount),0)
    FROM hamal_ledger
    WHERE DATE(created_at)=CURDATE()
    """)
    hamal_payments = cur.fetchone()[0] or 0

    # =========================
    # CASH IN FROM BILLS (REMOVE ADVANCE PART)
    # =========================
    cur.execute("""
    SELECT IFNULL(SUM(
        GREATEST(b.paid_amount - IFNULL(ab.advance_amount,0),0)
    ),0)
    FROM bills b
    LEFT JOIN advance_bookings ab
        ON b.advance_id = ab.id
    WHERE DATE(b.created_at)=CURDATE()
    AND b.upi_account='Cash'
    AND b.status IN ('DONE','BAKI')
    """)
    cash_bills = cur.fetchone()[0] or 0

    # =========================
    # CASH FROM CHILLAR
    # =========================
    cur.execute("""
    SELECT IFNULL(SUM(amount),0)
    FROM chillar_entries
    WHERE DATE(created_at)=CURDATE()
    AND payment_mode='Cash'
    """)
    chillar_cash = cur.fetchone()[0] or 0

    # =========================
    # TOTAL CASH RECEIVED
    # =========================
    cash_received = float(cash_bills) + float(chillar_cash) + float(advance_cash)

    total_collection = float(cash_received) + float(upi_collection)

    cur.execute("""
    SELECT IFNULL(SUM(amount),0)
    FROM expenses
    WHERE DATE(created_at)=CURDATE()
    """)
    other_expenses = cur.fetchone()[0] or 0

    today_expenses = (
            float(supplier_payments)
            + float(staff_payments)
            + float(hamal_payments)
            + float(other_expenses)
    )
    cur.execute("""
    SELECT IFNULL(SUM(amount),0)
    FROM supplier_payments
    WHERE payment_method='Cash'
    AND DATE(payment_date)=CURDATE()
    """)
    cash_supplier_payment = cur.fetchone()[0] or 0

    # =========================
    # FINAL CASH IN GALLA
    # =========================
    cash_collection = (
            float(cash_received)
            - float(cash_supplier_payment)
            - float(staff_payments)
            - float(hamal_payments)
    )    # =========================
    # 5️⃣ MARKET PENDING (Bills + Chillar)
    # =========================
    cur.execute("""
    SELECT 
        IFNULL(
            (SELECT SUM(balance_amount)
             FROM bills
             WHERE balance_amount > 0
             AND status IN ('DONE','BAKI')
            ),0
        )
        +
        IFNULL(
            (SELECT SUM(amount)
             FROM chillar_entries
             WHERE type='BAKI'
             AND status='BAKI'
            ),0
        )
    """)

    total_pending = cur.fetchone()[0] or 0

    cur.execute("""
        SELECT COUNT(*)
        FROM bills
        WHERE balance_amount>0
        AND status!='ESTIMATE'
    """)
    pending_bills = cur.fetchone()[0] or 0

    # =========================
    # 6️⃣ LOW STOCK
    # =========================
    cur.execute("""
        SELECT COUNT(*)
        FROM inventory
        WHERE stock_quantity <= min_stock_level
    """)
    low_stock_count = cur.fetchone()[0] or 0

    # =========================
    # 7️⃣ PROFIT (OWNER)
    # =========================
    net_profit = 0
    if session.get("role") == "owner":
        cur.execute("""
            SELECT IFNULL(SUM((bi.rate - bi.purchase_price) * bi.quantity),0)
            FROM bill_items bi
            JOIN bills b ON bi.bill_id = b.id
            WHERE DATE(b.created_at)=CURDATE()
            AND b.status IN ('DONE','BAKI')
            AND b.advance_id IS NULL
        """)
        net_profit = cur.fetchone()[0] or 0

    # =========================
    # 8️⃣ STAFF ACTIVITY (WITH EDIT COUNT)
    # =========================
    if session.get("role") == "owner":
        cur.execute("""
            SELECT 
                u.username,
                COUNT(DISTINCT b.id) AS bill_count,
                IFNULL(SUM(b.total_amount),0) AS total_sale,
                COUNT(r.id) AS edits
            FROM users u
            LEFT JOIN bills b 
                ON u.id = b.created_by 
                AND DATE(b.created_at)=CURDATE()
                AND b.status!='ESTIMATE'
            LEFT JOIN rate_edit_log r
                ON u.id = r.user_id
            WHERE u.role='staff'
            GROUP BY u.id,u.username
            ORDER BY u.username
        """)
    else:
        cur.execute("""
            SELECT 
                u.username,
                COUNT(DISTINCT b.id) AS bill_count,
                IFNULL(SUM(b.total_amount),0) AS total_sale,
                COUNT(r.id) AS edits
            FROM users u
            LEFT JOIN bills b 
                ON u.id = b.created_by 
                AND DATE(b.created_at)=CURDATE()
                AND b.status!='ESTIMATE'
            LEFT JOIN rate_edit_log r
                ON u.id = r.user_id
            WHERE u.id=%s
            GROUP BY u.id,u.username
        """, (session["user_id"],))

    rows = cur.fetchall()

    staff_data = []
    for r in rows:
        staff_data.append({
            "name": r[0],
            "bill_count": int(r[1] or 0),
            "total_sale": float(r[2] or 0),
            "edits": int(r[3] or 0)  # ✅ REAL EDIT COUNT
        })

    # =========================
    # 9️⃣ RATE EDIT HISTORY (OWNER)
    # =========================
    rate_edits = []
    if session.get("role") == "owner":
        cur.execute("""
            SELECT 
                u.username,
                i.product_name,
                r.old_price,
                r.new_price,
                r.changed_at
            FROM rate_edit_log r
            JOIN users u ON r.user_id = u.id
            JOIN inventory i ON r.product_id = i.id
            ORDER BY r.changed_at DESC
            LIMIT 10
        """)

        rows = cur.fetchall()

        for r in rows:
            rate_edits.append({
                "user": r[0],
                "product": r[1],
                "old": float(r[2]),
                "new": float(r[3]),
                "time": r[4].strftime("%d %b %H:%M")
            })
    cur.close()

    return render_template(
        "dashboard.html",
        username=session.get("username"),
        role=session.get("role"),
        total_sale=total_sale,
        upi_collection=upi_collection,
        cash_collection=cash_collection,
        today_expenses=today_expenses,
        net_profit=net_profit,
        total_pending=total_pending,
        pending_bills=pending_bills,
        low_stock_count=low_stock_count,
        staff_data=staff_data,
        rate_edits=rate_edits,  # 👈 ADD THIS
        total_collection=total_collection,
        upi_balance=upi_balance,
        current_date=date.today().strftime("%d %b %Y")
    )

@app.route("/profit-report")
@login_required
@owner_required
def profit_report():

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # ================= BASIC PROFITS =================

    cur.execute("""
        SELECT IFNULL(SUM((bi.rate-bi.purchase_price)*bi.quantity),0) AS profit
        FROM bill_items bi
        JOIN bills b ON bi.bill_id=b.id
        WHERE DATE(b.created_at)=CURDATE()
        AND b.status IN ('DONE','BAKI')
    """)
    today = cur.fetchone()["profit"]

    cur.execute("""
        SELECT IFNULL(SUM((bi.rate-bi.purchase_price)*bi.quantity),0) AS profit
        FROM bill_items bi
        JOIN bills b ON bi.bill_id=b.id
        WHERE DATE(b.created_at)=CURDATE()-INTERVAL 1 DAY
        AND b.status IN ('DONE','BAKI')
    """)
    yesterday = cur.fetchone()["profit"]

    cur.execute("""
        SELECT IFNULL(SUM((bi.rate-bi.purchase_price)*bi.quantity),0) AS profit
        FROM bill_items bi
        JOIN bills b ON bi.bill_id=b.id
        WHERE YEARWEEK(b.created_at,1)=YEARWEEK(CURDATE(),1)
        AND b.status IN ('DONE','BAKI')
    """)
    week = cur.fetchone()["profit"]

    cur.execute("""
        SELECT IFNULL(SUM((bi.rate-bi.purchase_price)*bi.quantity),0) AS profit
        FROM bill_items bi
        JOIN bills b ON bi.bill_id=b.id
        WHERE YEAR(b.created_at)=YEAR(CURDATE())
        AND MONTH(b.created_at)=MONTH(CURDATE())
        AND b.status IN ('DONE','BAKI')
    """)
    month = cur.fetchone()["profit"]

    cur.execute("""
        SELECT IFNULL(SUM((bi.rate-bi.purchase_price)*bi.quantity),0) AS profit
        FROM bill_items bi
        JOIN bills b ON bi.bill_id=b.id
        WHERE YEAR(b.created_at)=YEAR(CURDATE())
        AND b.status IN ('DONE','BAKI')
    """)
    year = cur.fetchone()["profit"]

    cur.execute("""
        SELECT IFNULL(SUM((bi.rate-bi.purchase_price)*bi.quantity),0) AS profit
        FROM bill_items bi
        JOIN bills b ON bi.bill_id=b.id
        WHERE b.status IN ('DONE','BAKI')
    """)
    all_time = cur.fetchone()["profit"]

    # ================= PRODUCT WISE PROFIT =================

    cur.execute("""
        SELECT 
            i.product_name,
            SUM((bi.rate-bi.purchase_price)*bi.quantity) AS profit
        FROM bill_items bi
        JOIN inventory i ON bi.product_id=i.id
        JOIN bills b ON bi.bill_id=b.id
        WHERE b.status IN ('DONE','BAKI')
        GROUP BY bi.product_id
        ORDER BY profit DESC
        LIMIT 10
    """)
    product_profit = cur.fetchall()

    # ================= STAFF WISE PROFIT =================

    cur.execute("""
        SELECT 
            u.name,
            SUM((bi.rate-bi.purchase_price)*bi.quantity) AS profit
        FROM bill_items bi
        JOIN bills b ON bi.bill_id=b.id
        JOIN users u ON b.created_by=u.id
        WHERE b.status IN ('DONE','BAKI')
        GROUP BY u.id
        ORDER BY profit DESC
    """)
    staff_profit = cur.fetchall()

    # ================= MONTHLY TREND =================
    cur.execute("""
        SELECT 
            MONTH(b.created_at) AS month_no,
            DATE_FORMAT(MIN(b.created_at),'%b') AS month,
            SUM((bi.rate - bi.purchase_price) * bi.quantity) AS profit
        FROM bill_items bi
        JOIN bills b ON bi.bill_id = b.id
        WHERE YEAR(b.created_at) = YEAR(CURDATE())
        AND b.status IN ('DONE','BAKI')
        GROUP BY MONTH(b.created_at)
        ORDER BY month_no
    """)
    monthly_profit = cur.fetchall()

    cur.close()

    return render_template(
        "profit_report.html",
        today=today,
        yesterday=yesterday,
        week=week,
        month=month,
        year=year,
        all_time=all_time,
        product_profit=product_profit,
        staff_profit=staff_profit,
        monthly_profit=monthly_profit
    )
@app.route("/chillar-bikri")
@login_required
def chillar_bikri():

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute("""
        SELECT * FROM chillar_entries
        WHERE type='BIKRI'
        ORDER BY created_at DESC
        LIMIT 20
    """)
    entries = cur.fetchall()
    cur.close()

    return render_template(
        "chillar_entry.html",
        mode="BIKRI",
        entries=entries
    )


@app.route("/chillar-baki")
@login_required
def chillar_baki():

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cur.execute("""
        SELECT 
            customer_name,
            customer_mobile,
            SUM(amount) AS total_baki
        FROM chillar_entries
        WHERE type='BAKI'
          AND status='BAKI'   -- ✅ only unpaid baki
        GROUP BY customer_name, customer_mobile
        HAVING SUM(amount) > 0
        ORDER BY total_baki DESC
    """)

    customers = cur.fetchall()

    cur.execute("SELECT id,name FROM users WHERE role='staff'")
    staff = cur.fetchall()
    cur.close()

    return render_template(
        "chillar_baki.html",
        customers=customers,
        staff=staff
    )

@app.route("/receive-chillar", methods=["POST"])
@login_required
def receive_chillar():

    name = request.form.get("customer_name")
    mobile = request.form.get("customer_mobile")
    amount_paid = float(request.form.get("amount_paid") or 0)
    payment_mode = request.form.get("payment_mode")

    # ✅ NEW: UPI receiver info
    upi_account = request.form.get("upi_account") if payment_mode == "UPI" else None
    staff_id = request.form.get("staff_id") if upi_account == "Staff" else None

    cur = mysql.connection.cursor()

    # ===== GET TOTAL BAKI ROWS =====
    cur.execute("""
        SELECT id, amount
        FROM chillar_entries
        WHERE customer_name=%s
          AND customer_mobile=%s
          AND type='BAKI'
          AND status='BAKI'
        ORDER BY created_at ASC
    """, (name, mobile))

    baki_rows = cur.fetchall()
    remaining = amount_paid

    # ===== CLEAR BAKI =====
    for row in baki_rows:
        if remaining <= 0:
            break

        baki_id = row[0]
        baki_amt = float(row[1])

        if remaining >= baki_amt:
            # full clear
            cur.execute("""
                UPDATE chillar_entries
                SET status='PAID',
                    received_at=NOW(),
                    payment_mode=%s,
                    upi_account=%s,
                    upi_staff_id=%s
                WHERE id=%s
            """, (payment_mode, upi_account, staff_id, baki_id))

            remaining -= baki_amt

        else:
            # partial clear
            new_amt = baki_amt - remaining

            cur.execute("""
                UPDATE chillar_entries
                SET amount=%s
                WHERE id=%s
            """, (new_amt, baki_id))

            # record received part as BIKRI
            cur.execute("""
                INSERT INTO chillar_entries
                (customer_name, customer_mobile,
                 amount, payment_mode,
                 upi_account, upi_staff_id,
                 type, status, created_at)
                VALUES (%s,%s,%s,%s,%s,%s,'BIKRI','PAID',NOW())
            """, (name, mobile, remaining, payment_mode, upi_account, staff_id))

            remaining = 0

    # ===== EXTRA PAYMENT → NEW BIKRI =====
    if remaining > 0:
        cur.execute("""
            INSERT INTO chillar_entries
            (customer_name, customer_mobile,
             amount, payment_mode,
             upi_account, upi_staff_id,
             type, status, created_at)
            VALUES (%s,%s,%s,%s,%s,%s,'BIKRI','PAID',NOW())
        """, (name, mobile, remaining, payment_mode, upi_account, staff_id))

    mysql.connection.commit()
    cur.close()

    return redirect(url_for("chillar_baki"))

@app.route("/save-chillar-bikri", methods=["POST"])
@login_required
def save_chillar_bikri():

    name = request.form.get("customer_name")
    mobile = request.form.get("customer_mobile")
    pid = request.form.get("product_id")
    qty = float(request.form.get("qty") or 0)
    rate = float(request.form.get("rate") or 0)
    amount = float(request.form.get("amount") or 0)
    payment = request.form.get("payment_mode")

    cur = mysql.connection.cursor()

    cur.execute("""
        INSERT INTO chillar_entries
        (customer_name, customer_mobile,
         product_id, qty, rate, amount,
         payment_mode, type, created_at)
        VALUES (%s,%s,%s,%s,%s,%s,%s,'BIKRI',NOW())
    """,(name,mobile,pid,qty,rate,amount,payment))

    # reduce stock
    cur.execute("""
        UPDATE inventory
        SET stock_quantity = stock_quantity - %s
        WHERE id=%s
    """,(qty,pid))

    mysql.connection.commit()
    cur.close()

    return jsonify({"success":True})

@app.route("/save-chillar-multi", methods=["POST"])
@login_required
def save_chillar_multi():

    name = request.form.get("customer_name")
    mobile = request.form.get("customer_mobile")
    payment = request.form.get("payment_mode")
    total = float(request.form.get("total") or 0)

    pids = request.form.getlist("pid[]")
    qtys = request.form.getlist("qty[]")
    rates = request.form.getlist("rate[]")

    # determine type + status
    if payment == "Pending":
        ctype = "BAKI"
        status = "BAKI"
    else:
        ctype = "BIKRI"
        status = "PAID"

    cur = mysql.connection.cursor()

    # insert chillar entry
    cur.execute("""
        INSERT INTO chillar_entries
        (customer_name, customer_mobile, amount,
         payment_mode, type, status, created_at)
        VALUES (%s,%s,%s,%s,%s,%s,NOW())
    """,(name, mobile, total, payment, ctype, status))

    chillar_id = cur.lastrowid

    # insert items
    for i in range(len(pids)):
        pid = pids[i]
        if not pid:
            continue

        q = float(qtys[i] or 0)
        r = float(rates[i] or 0)

        cur.execute("""
            INSERT INTO chillar_items
            (chillar_id, product_id, qty, rate)
            VALUES (%s,%s,%s,%s)
        """,(chillar_id, pid, q, r))

        # reduce stock only if paid
        if status == "PAID":
            cur.execute("""
                UPDATE inventory
                SET stock_quantity = stock_quantity - %s
                WHERE id=%s
            """,(q, pid))

    mysql.connection.commit()
    cur.close()

    return jsonify({"success":True})

@app.route("/api/product/<int:product_id>")
@login_required
def get_product(product_id):
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT id, name, selling_price, purchase_price
        FROM inventory
        WHERE id=%s
    """, (product_id,))
    item = cur.fetchone()
    cur.close()

    if item:
        return {
            "id": item[0],
            "name": item[1],
            "selling_price": float(item[2]),
            "purchase_price": float(item[3]),
            "profit": float(item[2] - item[3])
        }
    return {}

@app.route("/add-product-fast", methods=["POST"])
@login_required
def add_product_fast():
    data = request.get_json()
    name = data.get("name")

    cur = mysql.connection.cursor()
    cur.execute("""
        INSERT INTO inventory
        (product_name, stock_quantity, purchase_price, selling_price)
        VALUES (%s,0,0,0)
    """,(name,))

    new_id = cur.lastrowid
    mysql.connection.commit()
    cur.close()

    return jsonify({"id": new_id, "name": name})

@app.route("/inventory", methods=["GET", "POST"])
@login_required
def inventory():

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    if request.method == "POST":

        import json

        products_json = request.form.get("products_json")

        # ===== HANDLE MULTIPLE PRODUCTS =====
        if products_json:

            products = json.loads(products_json)

            for p in products:

                name = (p.get("name") or "").strip().lower()
                unit = p.get("unit")

                purchase_price = float(p.get("purchase") or 0)
                selling_price = float(p.get("selling") or 0)

                stock = float(p.get("qty") or 0)
                min_stock = float(request.form.get("min_stock") or 0)

                supplier_id = p.get("supplier_id")
                supplier_name = p.get("supplier")

                supplier_id = get_or_create_supplier(cur, supplier_id, supplier_name)

                stock_type = request.form.get("stock_type") or "REGULAR"
                gst_rate = float(request.form.get("gst_rate") or 0)

                taxable_value = purchase_price * stock
                gst_amount = (taxable_value * gst_rate) / 100

                cgst = gst_amount / 2
                sgst = gst_amount / 2

                total_amount = taxable_value + gst_amount

                # check duplicate product
                cur.execute(
                    "SELECT id FROM inventory WHERE LOWER(product_name)=%s AND stock_type=%s",
                    (name, stock_type)
                )
                exists = cur.fetchone()

                if exists:

                    # UPDATE STOCK IF PRODUCT EXISTS
                    cur.execute("""
                        UPDATE inventory
                        SET stock_quantity = stock_quantity + %s,
                            purchase_price=%s,
                            selling_price=%s,
                            supplier_id=%s,
                            rate_updated_at=NOW()
                        WHERE id=%s
                    """, (
                        stock,
                        purchase_price,
                        selling_price,
                        supplier_id,
                        exists["id"]
                    ))

                else:

                    # INSERT INVENTORY
                    cur.execute("""
                        INSERT INTO inventory
                        (product_name, unit, purchase_price,
                         selling_price, stock_quantity,
                         min_stock_level, stock_type,
                         supplier_id, gst_rate,
                         taxable_value, cgst, sgst,
                         total_amount, purchase_date,
                         rate_updated_at)

                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),NOW())
                    """, (
                        name, unit, purchase_price, selling_price,
                        stock, min_stock, stock_type,
                        supplier_id, gst_rate,
                        taxable_value, cgst, sgst, total_amount
                    ))

                # SUPPLIER LEDGER ENTRY
                purchase_amount = purchase_price * stock

                cur.execute("""
                    INSERT INTO supplier_ledger
                    (
                    supplier_id,
                    bill_no,
                    credit_amount,
                    debit_amount,
                    balance_amount,
                    transaction_date
                    )
                    VALUES (%s,%s,%s,%s,%s,NOW())
                """, (
                    supplier_id,
                    name,
                    purchase_amount,
                    0,
                    purchase_amount
                ))

            mysql.connection.commit()

            flash("Products Added Successfully!", "success")
            return redirect(url_for("inventory"))

        # ===== ORIGINAL SINGLE PRODUCT LOGIC (UNCHANGED) =====

        name = request.form.get("product_name", "").strip().lower()
        unit = request.form.get("unit")

        purchase_price = float(request.form.get("purchase_price") or 0)
        selling_price = float(request.form.get("selling_price") or 0)

        stock_type = request.form.get("stock_type") or "REGULAR"

        stock = float(request.form.get("stock") or 0)
        min_stock = float(request.form.get("min_stock") or 0)

        supplier_id = request.form.get("supplier_id")
        supplier_name = request.form.get("supplier_name")

        supplier_id = get_or_create_supplier(cur, supplier_id, supplier_name)

        gst_rate = float(request.form.get("gst_rate") or 0)

        taxable_value = purchase_price * stock
        tax_type = request.form.get("tax_type", "INTRA")

        tax_data = calculate_gst_split(taxable_value, gst_rate, tax_type)

        gst_amount = tax_data["gst_amount"]
        cgst = tax_data["cgst"]
        sgst = tax_data["sgst"]
        igst = tax_data["igst"]

        total_amount = taxable_value + gst_amount


        # check duplicate product
        cur.execute(
            "SELECT id FROM inventory WHERE LOWER(product_name)=%s AND stock_type=%s",
            (name, stock_type)
        )
        exists = cur.fetchone()

        if exists:

            flash("Product already exists for this stock type!", "warning")

        else:

            # INSERT INVENTORY
            cur.execute("""
                    INSERT INTO inventory
                    (product_name, unit, purchase_price,
                     selling_price, stock_quantity,
                     min_stock_level, stock_type,
                     supplier_id, gst_rate,
                     taxable_value, cgst, sgst, igst,
                     total_amount, tax_type, purchase_date,
                     rate_updated_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),NOW())
            """, (
    name, unit, purchase_price, selling_price,
    stock, min_stock, stock_type,
    supplier_id, gst_rate,
    taxable_value, cgst, sgst, igst,
    total_amount, tax_type
))

            purchase_amount = purchase_price * stock

            cur.execute("""
                INSERT INTO supplier_ledger
                (
                supplier_id,
                bill_no,
                credit_amount,
                debit_amount,
                balance_amount,
                transaction_date
                )
                VALUES (%s,%s,%s,%s,%s,NOW())
            """, (
                supplier_id,
                name,
                purchase_amount,
                0,
                purchase_amount
            ))

            mysql.connection.commit()

            flash("Product Added Successfully!", "success")

    # LOAD PRODUCTS
    cur.execute("""
        SELECT i.*, s.name as supplier_name
        FROM inventory i
        LEFT JOIN suppliers s ON i.supplier_id = s.id
        ORDER BY i.stock_type, i.product_name ASC
    """)

    items = cur.fetchall()

    cur.execute("SELECT id,name FROM suppliers ORDER BY name")
    suppliers = cur.fetchall()

    product_id = request.args.get("product_id")
    product = None

    if product_id:
        cur.execute("SELECT * FROM inventory WHERE id=%s", (product_id,))
        product = cur.fetchone()
    cur.close()

    return render_template(
        "inventory.html",
        items=items,
        suppliers=suppliers,
        product=product,
        role=session.get("role")
    )

@app.route("/delete-rate-edit/<int:edit_id>")
@login_required
@owner_required
def delete_rate_edit(edit_id):
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM rate_edit_log WHERE id=%s", (edit_id,))
    mysql.connection.commit()
    cur.close()

    flash("Rate edit deleted", "info")
    return redirect(url_for("rate_update_page"))

@app.route("/edit-rate-edit/<int:edit_id>", methods=["POST"])
@login_required
@owner_required
def edit_rate_edit(edit_id):
    new_price = request.form.get("new_price")

    cur = mysql.connection.cursor()

    # update log
    cur.execute("""
        UPDATE rate_edit_log
        SET new_price=%s
        WHERE id=%s
    """, (new_price, edit_id))

    mysql.connection.commit()
    cur.close()

    flash("Edit updated", "success")
    return redirect(url_for("rate_update_page"))

@app.route("/delete-product/<int:product_id>")
@login_required
@owner_required
def delete_product(product_id):
    cur = mysql.connection.cursor()

    cur.execute("SELECT id FROM inventory WHERE id=%s",(product_id,))
    if not cur.fetchone():
        flash("Product not found","danger")
        return redirect(url_for("inventory"))

    # delete dependent records
    cur.execute("DELETE FROM rate_history WHERE product_id=%s",(product_id,))

    # delete product
    cur.execute("DELETE FROM inventory WHERE id=%s",(product_id,))

    mysql.connection.commit()
    cur.close()

    flash("Product deleted successfully","success")
    return redirect(url_for("inventory"))

@app.route("/update-rate/<int:product_id>", methods=["POST"])
@login_required
@owner_required
def update_rate(product_id):

    cur = mysql.connection.cursor()

    new_name = request.form.get("product_name")
    new_stock = float(request.form.get("stock_quantity") or 0)

    new_purchase = float(request.form.get("purchase_price") or 0)
    new_selling = float(request.form.get("selling_price") or 0)
    # Get old prices
    cur.execute("""
        SELECT purchase_price, selling_price
        FROM inventory
        WHERE id=%s
    """, (product_id,))

    old = cur.fetchone()

    if not old:
        flash("Product not found!", "danger")
        return redirect(url_for("inventory"))

    old_purchase = old[0]
    old_selling = old[1]

    # Update inventory
    cur.execute("""
        UPDATE inventory
        SET product_name=%s,
            purchase_price=%s,
            selling_price=%s,
            stock_quantity=%s,
            rate_updated_at=NOW()
        WHERE id=%s
    """, (new_name, new_purchase, new_selling, new_stock, product_id))

    # Save history
    cur.execute("""
        INSERT INTO rate_history
        (product_id, old_purchase_price, old_selling_price,
         new_purchase_price, new_selling_price, changed_by)
        VALUES (%s,%s,%s,%s,%s,%s)
    """, (
        product_id,
        old_purchase,
        old_selling,
        new_purchase,
        new_selling,
        session["user_id"]
    ))

    mysql.connection.commit()
    cur.close()

    flash("Rate Updated Successfully!", "success")
    return redirect(url_for("inventory"))

@app.route("/rate-history/<int:product_id>")
@login_required
@owner_required
def rate_history(product_id):

    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT rh.old_selling_price,
               rh.new_selling_price,
               u.username,
               rh.changed_at
        FROM rate_history rh
        JOIN users u ON rh.changed_by = u.id
        WHERE rh.product_id=%s
        ORDER BY rh.changed_at DESC
    """, (product_id,))

    history = cur.fetchall()
    cur.close()

    return render_template("rate_history.html", history=history)


@app.route("/api/dashboard-data")
@login_required
def dashboard_data(upi_today=None):

    cur = mysql.connection.cursor()

    # Total Sale Today
    cur.execute("""
    SELECT IFNULL(SUM(total_amount),0)
    FROM bills
    WHERE DATE(created_at)=CURDATE()
    AND status!='ESTIMATE'
    """)
    total_sale = cur.fetchone()[0]

    # UPI Today
    cur.execute("""
        SELECT IFNULL(SUM(paid_amount),0)
        FROM bills
        WHERE upi_account IN ('Owner','Staff')
        AND DATE(created_at)=CURDATE()
        AND status!='ESTIMATE'
    """)
    upi_collection = cur.fetchone()[0] or 0

    # Expenses Today
    cur.execute("SELECT IFNULL(SUM(amount),0) FROM expenses WHERE DATE(created_at)=CURDATE()")
    expense_today = cur.fetchone()[0]

    # Pending
    cur.execute("""
    SELECT IFNULL(SUM(balance_amount),0)
    FROM bills
    WHERE balance_amount>0
    AND status!='ESTIMATE'
    """)
    total_pending = cur.fetchone()[0]

    # Profit (only owner)
    total_profit = 0
    if session.get("role") == "owner":
        cur.execute("""
            SELECT IFNULL(SUM((bi.rate - bi.purchase_price)*bi.quantity),0)
            FROM bill_items bi
            JOIN bills b ON bi.bill_id=b.id
WHERE DATE(b.created_at)=CURDATE()
AND b.status!='ESTIMATE'
        """)
        total_profit = cur.fetchone()[0]

    cur.close()

    return {
        "total_sale": float(total_sale),
        "upi_today": float(upi_today),
        "expense_today": float(expense_today),
        "total_pending": float(total_pending),
        "total_profit": float(total_profit),
        "galla_cash": float(total_sale - upi_today - expense_today)
    }



@app.route('/advance-booking', methods=['GET', 'POST'])
def advance_booking():

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    if request.method == "POST":
        try:
            customer_name = request.form.get("customer_name")
            mobile = request.form.get("mobile")

            advance_amount = float(request.form.get("advance_amount") or 0)
            locked_rate = float(request.form.get("locked_rate") or 0)
            delivery_date = request.form.get("delivery_date")

            payment_mode = request.form.get("payment_mode")
            upi_id = request.form.get("upi_id")

            # ================= CHECK IF CUSTOMER EXISTS =================

            cursor.execute(
                "SELECT id FROM customers WHERE phone=%s",
                (mobile,)
            )

            existing = cursor.fetchone()

            if existing:
                # use existing customer
                customer_id = existing["id"]

            else:
                # create new customer
                cursor.execute("""
                    INSERT INTO customers (name, phone)
                    VALUES (%s,%s)
                """, (customer_name, mobile))

                customer_id = cursor.lastrowid


            # Create booking
            cursor.execute("""
            INSERT INTO advance_bookings
            (customer_id, advance_amount, locked_rate, delivery_date, payment_mode, upi_id, status, created_at)
            VALUES (%s,%s,%s,%s,%s,%s,'ACTIVE',NOW())
            """, (
                customer_id,
                advance_amount,
                locked_rate,
                delivery_date,
                payment_mode,
                upi_id
            ))

            booking_id = cursor.lastrowid

            # Insert items
            product_ids = request.form.getlist("product_id[]")
            product_names = request.form.getlist("product_name[]")
            qtys = request.form.getlist("qty[]")
            rates = request.form.getlist("rate[]")

            for i in range(len(product_names)):

                pname = product_names[i].strip()

                if pname == "":
                    continue

                pid = product_ids[i] if product_ids[i] else None
                qty = float(qtys[i] or 0)
                rate = float(rates[i] or 0)
                total = qty * rate

                cursor.execute("""
                INSERT INTO advance_booking_items
                (booking_id, product_id, product_name, qty, rate, total)
                VALUES (%s,%s,%s,%s,%s,%s)
                """, (
                    booking_id,
                    pid,
                    pname,
                    qty,
                    rate,
                    total
                ))
            mysql.connection.commit()

            flash("Advance Booking Saved Successfully!", "success")
            return redirect(url_for('advance_booking'))

        except Exception as e:
            mysql.connection.rollback()
            print("ERROR:", e)
            flash("Error saving booking!", "danger")

    # ================= FETCH BOOKINGS =================
    cursor.execute("""
    SELECT 
    ab.id,
    ab.advance_amount,
    ab.delivery_date,
    ab.status,
    ab.payment_mode,
    c.name as customer_name,
    c.phone
    FROM advance_bookings ab
    JOIN customers c ON ab.customer_id = c.id
    ORDER BY ab.id DESC
    """)
    bookings = cursor.fetchall()

    # ================= FETCH ITEMS =================
    cursor.execute("""
    SELECT booking_id, product_name, qty
    FROM advance_booking_items
    ORDER BY booking_id DESC
    """)
    items = cursor.fetchall()

    # ================= ATTACH ITEMS =================

    booking_map = {}

    for b in bookings:
        b["booking_items"] = []
        booking_map[b["id"]] = b

    for item in items:

        if item["booking_id"] in booking_map:
            booking_map[item["booking_id"]]["booking_items"].append(item)

    return render_template(
        "advance_booking.html",
        bookings=bookings
    )

@app.route("/generate-bill/<int:booking_id>")
def generate_bill(booking_id):

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cur.execute("""
    SELECT 
    ab.*,
    c.name,
    c.phone
    FROM advance_bookings ab
    JOIN customers c ON ab.customer_id = c.id
    WHERE ab.id=%s
    """,(booking_id,))

    booking = cur.fetchone()

    if not booking:
        return "Advance booking not found"

    cur.execute("""
    SELECT product_id, product_name, qty, rate
    FROM advance_booking_items
    WHERE booking_id=%s
    """,(booking_id,))

    advance_items = cur.fetchall()

    cur.execute("SELECT id,name FROM hamals")
    hamals = cur.fetchall()

    return render_template(
        "billing.html",
        advance_booking=booking,
        advance_items=advance_items,
        hamals=hamals,
        items=[],
        bill=None
    )

@app.route("/advance-receipt/<int:id>")
@login_required
def advance_receipt(id):

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Booking
    cur.execute("""
    SELECT 
        ab.*,
        c.name AS customer_name,
        c.phone
    FROM advance_bookings ab
    LEFT JOIN customers c ON ab.customer_id = c.id
    WHERE ab.id=%s
    """, (id,))
    booking = cur.fetchone()

    if not booking:
        return "Advance not found"

    # Items
    cur.execute("SELECT * FROM advance_booking_items WHERE booking_id=%s", (id,))
    items = cur.fetchall()

    cur.close()

    return render_template(
        "advance_receipt.html",
        booking=booking,
        items=items
    )

@app.route("/delete-advance/<int:id>")
@login_required
def delete_advance(id):

    cur = mysql.connection.cursor()

    cur.execute("DELETE FROM advance_booking_items WHERE booking_id=%s", (id,))
    cur.execute("DELETE FROM advance_bookings WHERE id=%s", (id,))

    mysql.connection.commit()
    cur.close()

    return redirect("/advance-booking")

@app.route("/api/search-product")
@login_required
def search_product():

    q = request.args.get("q","").strip()

    bill_type = request.args.get("type","REGULAR").upper()

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cur.execute("""
        SELECT 
            id,
            product_name,
            selling_price AS selling_rate,
            purchase_price,
            stock_quantity,
            stock_type
        FROM inventory
        WHERE product_name LIKE %s
        AND stock_type = %s
        ORDER BY product_name
        LIMIT 10
    """, (f"%{q}%", bill_type))

    data = cur.fetchall()

    cur.close()

    return jsonify(data)

@app.route("/api/global-search")
@login_required
def global_search():
    q = request.args.get("q", "").strip()

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    results = []

    # ================= PRODUCTS =================
    cur.execute("""
        SELECT id,
               product_name AS name,
               'product' AS type
        FROM inventory
        WHERE product_name LIKE %s
        LIMIT 5
    """, (f"%{q}%",))
    results += cur.fetchall()

    # ================= CUSTOMERS =================
    cur.execute("""
        SELECT 
            MIN(id) AS id,
            customer_name AS name,
            'customer' AS type
        FROM bills
        WHERE customer_name LIKE %s
        GROUP BY customer_name
        LIMIT 5
    """, (f"%{q}%",))
    results += cur.fetchall()

    # ================= STAFF =================
    cur.execute("""
        SELECT id,
               name,
               'staff' AS type
        FROM users
        WHERE name LIKE %s
        LIMIT 5
    """, (f"%{q}%",))
    results += cur.fetchall()

    # ================= SUPPLIERS =================
    cur.execute("""
        SELECT id,
               name,
               'supplier' AS type
        FROM suppliers
        WHERE name LIKE %s
        LIMIT 5
    """, (f"%{q}%",))
    results += cur.fetchall()
    # ================= BILLS =================
    cur.execute("""
        SELECT id,
               CONCAT('Bill #', id, ' - ', customer_name) AS name,
               'bill' AS type
        FROM bills
        WHERE customer_name LIKE %s
           OR id LIKE %s
        ORDER BY id DESC
        LIMIT 5
    """, (f"%{q}%", f"%{q}%"))

    results += cur.fetchall()

    cur.close()

    # ================= ERP PAGES =================
    pages = [
        {"name": "Dashboard", "type": "page", "url": "/dashboard"},
        {"name": "Users", "type": "page", "url": "/users"},
        {"name": "Inventory", "type": "page", "url": "/inventory"},
        {"name": "Suppliers", "type": "page", "url": "/suppliers"},
        {"name": "Cash Tally", "type": "page", "url": "/cash-tally"},
        {"name": "UPI Report", "type": "page", "url": "/upi-report"},
        {"name": "Daily Report", "type": "page", "url": "/daily-report"},
        {"name": "Pending Bills", "type": "page", "url": "/pending-bills"},
        {"name": "Credit Ledger", "type": "page", "url": "/credit-ledger"},
    ]

    for p in pages:
        if q.lower() in p["name"].lower():
            results.append(p)

    return jsonify(results)


@app.route("/save-bill", methods=["POST"])
@login_required
def save_bill():
    cur = mysql.connection.cursor()

    try:
        # ================= BASIC INFO =================
        edit_id = request.form.get("edit_bill_id")
        advance_id = request.form.get("advance_id") or None
        customer = request.form.get("customer_name")
        mobile = request.form.get("customer_mobile")
        buyer_gstin = request.form.get("buyer_gstin")
        status = request.form.get("status")

        bill_type = request.form.get("bill_type") or "Regular"

        payment_mode = request.form.get("upi_account")
        upi_account = request.form.get("upi_account")

        def safe_float(v):
            try:
                if v in ("", "undefined", None):
                    return 0
                return float(v)
            except:
                return 0

        paid = safe_float(request.form.get("paid_amount"))
        # ================= HANDLE ADVANCE =================
        advance_amount = 0

        if advance_id:
            cur.execute("""
                SELECT advance_amount
                FROM advance_bookings
                WHERE id=%s
            """, (advance_id,))

            adv = cur.fetchone()

            if adv:
                advance_amount = float(adv[0] or 0)
        # ================= INCLUDE ADVANCE PAYMENT =================
        advance_amount = 0

        if advance_id:
            cur.execute("""
                SELECT advance_amount
                FROM advance_bookings
                WHERE id=%s
            """, (advance_id,))

            adv = cur.fetchone()

            if adv:
                advance_amount = float(adv[0] or 0)

        # total paid including advance
        total_paid = paid + advance_amount
        hamali_val = safe_float(request.form.get("hamali"))
        bhada_val = safe_float(request.form.get("bhada"))
        hamal_id = request.form.get("hamal_id")

        upi_id = request.form.get("upi_id")
        upi_holder = request.form.get("upi_holder")

        gst_rate = float(request.form.get("gst_rate") or 0)

        names = request.form.getlist("pName[]")
        product_ids = request.form.getlist("pId[]")
        qtys = request.form.getlist("pQty[]")
        rates = request.form.getlist("pRate[]")
        purchase_rates = request.form.getlist("pPurchaseRate[]")
        units = request.form.getlist("pUnit[]")


        # ================= ITEMS PROCESS =================
        total_amt = 0
        item_data = []

        for i in range(len(names)):

            product_name = names[i]

            if product_name.strip() == "":
                continue

            pid = product_ids[i] if product_ids[i] else None

            pid = None
            if i < len(product_ids):
                if product_ids[i] not in ("", "undefined", None):
                    pid = product_ids[i]

            def safe_float(v):
                try:
                    if v in ("", "undefined", None):
                        return 0
                    return float(v)
                except:
                    return 0

            q = safe_float(qtys[i] if i < len(qtys) else 0)
            r = safe_float(rates[i] if i < len(rates) else 0)
            pr = safe_float(purchase_rates[i] if i < len(purchase_rates) else 0)

            if status != "ESTIMATE" and pid:
                cur.execute(
                    "SELECT stock_quantity FROM inventory WHERE id=%s",
                    (pid,)
                )

                res = cur.fetchone()
                current_stock = float(res[0]) if res else 0

            total_amt += (q * r)
            item_data.append((pid, q, r, pr))

        # ================= GST =================
        grand_total = total_amt + hamali_val + bhada_val
        # ================= ESTIMATE SAVE =================
        if bill_type == "ESTIMATE":

            cur.execute("""
            INSERT INTO estimates
            (customer_name, customer_mobile, total_amount, created_by, created_at)
            VALUES (%s,%s,%s,%s,NOW())
            """, (
                customer,
                mobile,
                grand_total,
                session['user_id']
            ))

            est_id = cur.lastrowid

            for i in range(len(names)):

                pname = names[i]
                if pname.strip() == "":
                    continue

                q = float(qtys[i] or 0)
                r = float(rates[i] or 0)

                cur.execute("""
                INSERT INTO estimate_items
                (estimate_id, product_name, quantity, rate, amount)
                VALUES (%s,%s,%s,%s,%s)
                """, (
                    est_id,
                    pname,
                    q,
                    r,
                    q * r
                ))

            mysql.connection.commit()

            return jsonify({
                "success": True,
                "bill_id": est_id
            })
        tax_type = request.form.get("tax_type", "INTRA").strip().upper()

        gst_amount = 0
        cgst_amount = 0
        sgst_amount = 0
        igst_amount = 0
        taxable_value = grand_total

        if bill_type == "GST" and gst_rate > 0:
            taxable_value = float(grand_total) / (1 + float(gst_rate) / 100)
            gst_amount = grand_total - taxable_value

            if tax_type == "INTER":
                igst_amount = gst_amount
                cgst_amount = 0
                sgst_amount = 0
            else:
                cgst_amount = gst_amount / 2
                sgst_amount = gst_amount / 2
                igst_amount = 0
        # ================= GET ADVANCE AMOUNT =================
        advance_amount = 0

        if advance_id:
            cur.execute("""
                SELECT advance_amount
                FROM advance_bookings
                WHERE id=%s
            """, (advance_id,))

            adv = cur.fetchone()

            if adv:
                advance_amount = float(adv[0] or 0)

        # ================= CALCULATE BALANCE =================
        balance = max(float(grand_total) - float(advance_amount) - float(paid), 0)

        # ================= STATUS LOGIC =================

        # If user saved draft
        if status == "PENDING":
            pass

        else:

            if balance <= 0:
                status = "DONE"

            else:
                status = "BAKI"
        # ================= UPDATE EXISTING BILL =================
        if edit_id:
            bill_id = edit_id

            cur.execute(
                "SELECT product_id, quantity FROM bill_items WHERE bill_id=%s",
                (bill_id,)
            )
            old_items = cur.fetchall()

            for pid, q in old_items:

                if bill_type == "GST":
                    cur.execute(
                        "UPDATE gst_inventory SET stock_quantity = stock_quantity + %s WHERE id=%s",
                        (q, pid)
                    )
                else:
                    cur.execute(
                        "UPDATE inventory SET stock_quantity = stock_quantity + %s WHERE id=%s",
                        (q, pid)
                    )

            cur.execute("DELETE FROM bill_items WHERE bill_id=%s", (bill_id,))

            cur.execute("""
                UPDATE bills SET
                customer_name=%s,
                customer_mobile=%s,
                buyer_gstin=%s,
                total_amount=%s,
                paid_amount=%s,
                balance_amount=%s,
                payment_mode=%s,
                status=%s,
                bill_type=%s,
                gst_rate=%s,
                gst_amount=%s,
                taxable_value=%s,
                cgst_amount=%s,
                sgst_amount=%s,
                igst_amount=%s,
                tax_type=%s,
                bhada_amount=%s,
                upi_account=%s,
                upi_id=%s
                WHERE id=%s
            """, (
                customer,
                mobile,
                buyer_gstin,
                grand_total,
                paid,
                balance,
                payment_mode,
                status,
                bill_type,
                gst_rate,
                gst_amount,
                taxable_value,
                cgst_amount,
                sgst_amount,
                igst_amount,
                tax_type,
                bhada_val,
                upi_account,
                upi_id,
                bill_id
            ))
        else:
            staff_id = None

            if upi_account == "Staff":
                staff_id = request.form.get("staff_id")

            cur.execute("""
                INSERT INTO bills
                (customer_name, customer_mobile, buyer_gstin, total_amount, paid_amount,
                 balance_amount, payment_mode, status, bill_type,
                 gst_rate, gst_amount, taxable_value,
                 cgst_amount, sgst_amount, igst_amount, tax_type,
                 bhada_amount, upi_account, upi_id, upi_staff_id,
                 advance_id, created_by, created_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
            """, (
                customer,
                mobile,
                buyer_gstin,
                grand_total,
                paid,
                balance,
                payment_mode,
                status,
                bill_type,
                gst_rate,
                gst_amount,
                taxable_value,
                cgst_amount,
                sgst_amount,
                igst_amount,
                tax_type,
                bhada_val,
                upi_account,
                upi_id,
                staff_id,
                advance_id,
                session['user_id']
            ))
            bill_id = cur.lastrowid

        # ================= INSERT ITEMS =================
        for i in range(len(item_data)):

            pid, q, r, pr = item_data[i]
            unit = units[i] if i < len(units) else "pcs"

            product_name = names[i]

            cur.execute("""
            INSERT INTO bill_items
            (bill_id, product_id, product_name, quantity, unit, rate, purchase_price)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
            """, (bill_id, pid, product_name, q, unit, r, pr))

            if status != "ESTIMATE":
                table = "inventory"

                # 1️⃣ If product selected from inventory
                if pid:

                    cur.execute(f"""
                    UPDATE {table}
                    SET stock_quantity = stock_quantity - %s
                    WHERE id=%s
                    """, (q, pid))

                # 2️⃣ If product typed manually
                else:

                    pname = product_name.strip().lower()

                    cur.execute(f"""
                    SELECT id, stock_quantity
                    FROM {table}
                    WHERE LOWER(product_name)=%s
                    LIMIT 1
                    """, (pname,))

                    prod = cur.fetchone()

                    if prod:
                        new_stock = float(prod[1] or 0) - float(q)

                        cur.execute(f"""
                        UPDATE {table}
                        SET stock_quantity=%s
                        WHERE id=%s
                        """, (new_stock, prod[0]))

                    else:

                        cur.execute(f"""
                        INSERT INTO {table}
                        (product_name, unit, purchase_price, selling_price,
                         stock_quantity, min_stock_level, stock_type, purchase_date)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,NOW())
                        """, (
                            pname,
                            unit,
                            pr,
                            r,
                            -q,
                            0,
                            bill_type.upper()
                        ))
            # ================= CUSTOMER LEDGER =================

        if grand_total > 0:

            # 1️⃣ Bill entry (Debit)
            cur.execute("""
                 INSERT INTO customer_ledger
                 (bill_id, customer_name, debit, credit, balance, created_at)
                 VALUES (%s,%s,%s,%s,%s,NOW())
             """, (
                bill_id,
                customer,
                grand_total,
                0,
                balance
            ))

            # 2️⃣ Advance entry
            if advance_amount > 0:
                cur.execute("""
                     INSERT INTO customer_ledger
                     (bill_id, customer_name, debit, credit, balance, created_at)
                     VALUES (%s,%s,%s,%s,%s,NOW())
                 """, (
                    bill_id,
                    customer,
                    0,
                    advance_amount,
                    balance
                ))

            # 3️⃣ Paid entry
            if paid > 0:
                cur.execute("""
                     INSERT INTO customer_ledger
                     (bill_id, customer_name, debit, credit, balance, created_at)
                     VALUES (%s,%s,%s,%s,%s,NOW())
                 """, (
                    bill_id,
                    customer,
                    0,
                    paid,
                    balance
                ))
        # ================= HAMAL LEDGER ENTRY =================
        if hamal_id and hamali_val > 0:

            cur.execute("SELECT total_due FROM hamals WHERE id=%s", (hamal_id,))
            row = cur.fetchone()

            current_due = float(row[0] or 0)
            new_balance = current_due + hamali_val

            cur.execute("""
            INSERT INTO hamal_ledger
            (hamal_id, description, credit_amount, debit_amount, balance_amount, created_at, bill_id)
            VALUES (%s,%s,%s,%s,%s,NOW(),%s)
            """, (
                hamal_id,
                f"Bill #{bill_id} Hamali",
                hamali_val,
                0,
                new_balance,
                bill_id
            ))

            cur.execute("""
                UPDATE hamals
                SET total_due=%s
                WHERE id=%s
            """, (new_balance, hamal_id))

        # ================= BHADA TO HAMAL LEDGER =================
        if hamal_id and bhada_val > 0:
            cur.execute("SELECT total_due FROM hamals WHERE id=%s", (hamal_id,))
            row = cur.fetchone()

            current_due = float(row[0] or 0)
            new_balance = current_due + bhada_val

            cur.execute("""
                INSERT INTO hamal_ledger
                (hamal_id, description, credit_amount, debit_amount, balance_amount, created_at, bill_id)
                VALUES (%s,%s,%s,%s,%s,NOW(),%s)
            """, (
                hamal_id,
                f"Bhada - {customer}",
                bhada_val,
                0,
                new_balance,
                bill_id
            ))

            cur.execute("""
                UPDATE hamals
                SET total_due=%s
                WHERE id=%s
            """, (new_balance, hamal_id))
        # ================= MARK ADVANCE USED =================
        if advance_id:
            cur.execute("""
                UPDATE advance_bookings
                SET status='USED'
                WHERE id=%s
            """, (advance_id,))


        mysql.connection.commit()

        return jsonify({
            "success": True,
            "bill_id": bill_id
        })

    except Exception as e:
        mysql.connection.rollback()
        return jsonify({"success": False, "msg": str(e)})

    finally:
        cur.close()

@app.route("/bill-preview/<int:bill_id>")
@login_required
def bill_preview(bill_id):

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cur.execute("""
        SELECT b.*, bi.quantity, bi.rate, i.product_name
        FROM bills b
        JOIN bill_items bi ON b.id=bi.bill_id
        JOIN inventory i ON bi.product_id=i.id
        WHERE b.id=%s
    """,(bill_id,))

    bill = cur.fetchone()
    cur.close()

    return render_template("bill_preview.html", bill=bill)

from decimal import Decimal

@app.route("/daily-report")
@login_required
def daily_report():
    filter_type = request.args.get("type")
    from_date = request.args.get("from")
    to_date = request.args.get("to")

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # ================= DATE CONDITION =================
    if from_date and to_date:
        date_cond_b = f"DATE(b.created_at) BETWEEN '{from_date}' AND '{to_date}'"
        date_cond_c = f"DATE(c.created_at) BETWEEN '{from_date}' AND '{to_date}'"
        date_cond_e = f"DATE(e.created_at) BETWEEN '{from_date}' AND '{to_date}'"
        label = f"{from_date} → {to_date}"

    elif filter_type == "today":
        date_cond_b = "DATE(b.created_at)=CURDATE()"
        date_cond_c = "DATE(c.created_at)=CURDATE()"
        date_cond_e = "DATE(e.created_at)=CURDATE()"
        label = "Today"

    elif filter_type == "yesterday":
        date_cond_b = "DATE(b.created_at)=CURDATE()-INTERVAL 1 DAY"
        date_cond_c = "DATE(c.created_at)=CURDATE()-INTERVAL 1 DAY"
        date_cond_e = "DATE(e.created_at)=CURDATE()-INTERVAL 1 DAY"
        label = "Yesterday"

    elif filter_type == "week":
        date_cond_b = "YEARWEEK(b.created_at,1)=YEARWEEK(CURDATE(),1)"
        date_cond_c = "YEARWEEK(c.created_at,1)=YEARWEEK(CURDATE(),1)"
        date_cond_e = "YEARWEEK(e.created_at,1)=YEARWEEK(CURDATE(),1)"
        label = "This Week"

    elif filter_type == "month":
        date_cond_b = "YEAR(b.created_at)=YEAR(CURDATE()) AND MONTH(b.created_at)=MONTH(CURDATE())"
        date_cond_c = "YEAR(c.created_at)=YEAR(CURDATE()) AND MONTH(c.created_at)=MONTH(CURDATE())"
        date_cond_e = "YEAR(e.created_at)=YEAR(CURDATE()) AND MONTH(e.created_at)=MONTH(CURDATE())"
        label = "This Month"

    elif filter_type == "year":
        date_cond_b = "YEAR(b.created_at)=YEAR(CURDATE())"
        date_cond_c = "YEAR(c.created_at)=YEAR(CURDATE())"
        date_cond_e = "YEAR(e.created_at)=YEAR(CURDATE())"
        label = "This Year"

    else:
        date_cond_b = "DATE(b.created_at)=CURDATE()"
        date_cond_c = "DATE(c.created_at)=CURDATE()"
        date_cond_e = "DATE(e.created_at)=CURDATE()"
        label = "Today"

    # ================= DATE RANGE FOR EXTRA TABLES =================
    if not from_date:
        from_date = date.today().strftime("%Y-%m-%d")
    if not to_date:
        to_date = date.today().strftime("%Y-%m-%d")

    # ================= TOTAL SALE =================
    cur.execute(f"""
    SELECT 
        CASE 
            WHEN payment_mode IN ('Owner','Staff','UPI') THEN 'UPI'
            ELSE payment_mode
        END AS payment_mode,
        SUM(total_bill) AS total_bill,
        SUM(total_received) AS total_received
    FROM (
        SELECT b.payment_mode,
               SUM(b.total_amount) AS total_bill,
               SUM(b.paid_amount) AS total_received
        FROM bills b
        WHERE {date_cond_b}
        AND b.status IN ('DONE','BAKI')
        GROUP BY b.payment_mode

        UNION ALL

        SELECT c.payment_mode,
               SUM(c.amount) AS total_bill,
               SUM(c.amount) AS total_received
        FROM chillar_entries c
        WHERE {date_cond_c}
        AND c.type='BIKRI'
        GROUP BY c.payment_mode
    ) AS combined
    GROUP BY 
        CASE 
            WHEN payment_mode IN ('Owner','Staff','UPI') THEN 'UPI'
            ELSE payment_mode
        END
    """)

    sales_summary = cur.fetchall()

    # ================= EXPENSES =================
    cur.execute(f"""
    SELECT IFNULL(SUM(amount),0) AS total
    FROM expenses e
    WHERE {date_cond_e}
    """)
    expense_total = Decimal(cur.fetchone()['total'] or 0)

    # ================= SUPPLIER PAYMENTS =================
    cur.execute(f"""
    SELECT
        IFNULL(SUM(CASE WHEN payment_method='Cash' THEN amount ELSE 0 END),0) as cash_jawak,
        IFNULL(SUM(CASE WHEN payment_method='UPI' THEN amount ELSE 0 END),0) as upi_jawak,
        IFNULL(SUM(amount),0) as total
    FROM supplier_payments
    WHERE DATE(payment_date) BETWEEN '{from_date}' AND '{to_date}'
    """)

    supplier_data = cur.fetchone()

    cash_jawak = Decimal(supplier_data['cash_jawak'] or 0)
    upi_jawak = Decimal(supplier_data['upi_jawak'] or 0)
    supplier_payments = Decimal(supplier_data['total'] or 0)

    # ================= STAFF SALARY =================
    cur.execute(f"""
    SELECT IFNULL(SUM(amount),0) AS total
    FROM staff_payouts
    WHERE DATE(payout_date) BETWEEN '{from_date}' AND '{to_date}'
    """)
    staff_salary = Decimal(cur.fetchone()['total'] or 0)

    # ================= HAMAL PAYMENTS =================
    cur.execute(f"""
    SELECT IFNULL(SUM(debit_amount),0) AS total
    FROM hamal_ledger
    WHERE DATE(created_at) BETWEEN '{from_date}' AND '{to_date}'
    """)
    hamal_payments = Decimal(cur.fetchone()['total'] or 0)

    # FINAL EXPENSE
    total_expenses = (
            expense_total
            + supplier_payments
            + staff_salary
            + hamal_payments
    )

    # ================= CASH IN GALLA =================
    cur.execute(f"""
        SELECT 
            IFNULL(
                (SELECT SUM(b.paid_amount)
                 FROM bills b
                 WHERE b.payment_mode='Cash'
                 AND {date_cond_b}
                 AND b.status!='ESTIMATE'),0
            )
            +
            IFNULL(
                (SELECT SUM(c.amount)
                 FROM chillar_entries c
                 WHERE c.payment_mode='Cash'
                 AND {date_cond_c}
                 AND c.type IN ('BIKRI','RECEIPT')),0
            ) AS cash_in
    """)

    cash_received = Decimal(cur.fetchone()['cash_in'] or 0)

    net_galla_cash = cash_received - (
            expense_total
            + cash_jawak
            + staff_salary
            + hamal_payments
    )

    # ================= PROFIT =================
    cur.execute(f"""
        SELECT IFNULL(SUM((bi.rate - bi.purchase_price) * bi.quantity),0) AS profit
        FROM bill_items bi
        JOIN bills b ON bi.bill_id = b.id
        WHERE {date_cond_b}
        AND b.status IN ('DONE','BAKI')
    """)

    total_profit = Decimal(cur.fetchone()['profit'] or 0)

    net_cash_received = cash_received
    # ================= JAWAK HISTORY =================

    cur.execute(f"""
    SELECT amount, description, created_at
    FROM expenses
    WHERE DATE(created_at) BETWEEN '{from_date}' AND '{to_date}'
    ORDER BY created_at DESC
    LIMIT 20
    """)

    jawak_history = cur.fetchall()
    cur.close()

    return render_template(
        "daily_report.html",
        sales=sales_summary,
        expenses=total_expenses,
        galla=net_galla_cash,
        period_label=label,
        profit=total_profit,
        jawak_history=jawak_history,
        current_date=date.today()
    )

@app.route("/payment-history")
@login_required
def payment_history():

    mode = request.args.get("mode","Cash")
    page = int(request.args.get("page",1))
    per_page = 7
    offset = (page-1)*per_page

    from_date = request.args.get("from")
    to_date = request.args.get("to")
    customer = request.args.get("customer","").strip()

    where = ["payment_mode=%s"]
    params = [mode]

    if from_date:
        where.append("DATE(created_at)>=%s")
        params.append(from_date)

    if to_date:
        where.append("DATE(created_at)<=%s")
        params.append(to_date)

    if customer:
        where.append("customer_name LIKE %s")
        params.append(f"%{customer}%")

    where_sql = " AND ".join(where)

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # COUNT
    cur.execute(f"""
        SELECT COUNT(*) AS total
        FROM payments_view
        WHERE {where_sql}
    """, params)

    total_rows = cur.fetchone()["total"]
    total_pages = (total_rows//per_page)+(1 if total_rows%per_page else 0)

    # DATA
    cur.execute(f"""
        SELECT *
        FROM payments_view
        WHERE {where_sql}
        ORDER BY created_at DESC
        LIMIT %s OFFSET %s
    """, params+[per_page,offset])

    payments = cur.fetchall()

    # TOTAL AMOUNT
    cur.execute(f"""
        SELECT IFNULL(SUM(amount),0) AS total
        FROM payments_view
        WHERE {where_sql}
    """, params)

    total_amount = cur.fetchone()["total"]

    cur.close()

    return render_template(
        "payment_history.html",
        payments=payments,
        payment_mode=mode,
        total_rows=total_rows,
        total_pages=total_pages,
        page=page,
        total_amount=total_amount,
        period_label="Selected Period",
        export_url=request.url.replace("/payment-history","/export-payments")
    )


# GET supplier data (for modal edit)
@app.route("/get-supplier/<int:id>")
@login_required
def get_supplier(id):

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cur.execute("SELECT id,name,mobile,address FROM suppliers WHERE id=%s",(id,))
    supplier = cur.fetchone()

    cur.close()

    return jsonify(supplier)


# UPDATE supplier
@app.route("/update-supplier/<int:id>", methods=["POST"])
@login_required
def update_supplier(id):

    name = request.form.get("name")
    mobile = request.form.get("mobile")
    address = request.form.get("address")

    cur = mysql.connection.cursor()

    cur.execute("""
        UPDATE suppliers
        SET name=%s, mobile=%s, address=%s
        WHERE id=%s
    """,(name,mobile,address,id))

    mysql.connection.commit()
    cur.close()

    return jsonify({"success":True})

@app.route("/delete-supplier/<int:id>")
@login_required
def delete_supplier(id):
    cur = mysql.connection.cursor()

    # check ledger exists
    cur.execute("SELECT COUNT(*) FROM supplier_ledger WHERE supplier_id=%s", (id,))
    count = cur.fetchone()[0]

    if count > 0:
        # deactivate instead of delete
        cur.execute("UPDATE suppliers SET is_active=0 WHERE id=%s", (id,))
        mysql.connection.commit()
        flash("Supplier has ledger history. Deactivated instead.", "warning")
    else:
        cur.execute("DELETE FROM suppliers WHERE id=%s", (id,))
        mysql.connection.commit()
        flash("Supplier deleted", "success")

    cur.close()
    return redirect(url_for("suppliers"))

@app.route("/add-supplier", methods=["GET", "POST"])
@login_required
def add_supplier():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    if request.method == "POST":
        try:
            import json
            from datetime import date

            name = request.form.get("name", "").strip()
            mobile = request.form.get("mobile", "").strip()
            address = request.form.get("address", "").strip()
            bill_no = request.form.get("bill_no", "").strip()
            bill_date = request.form.get("bill_date") or date.today()

            products_json = request.form.get("products_json")

            if not name:
                flash("Supplier name is required", "danger")
                return redirect(url_for("add_supplier"))

            cur.execute("SELECT id FROM suppliers WHERE name=%s", (name,))
            supplier = cur.fetchone()

            if supplier:
                supplier_id = supplier["id"]
                cur.execute("""
                    UPDATE suppliers
                    SET mobile=%s, address=%s
                    WHERE id=%s
                """, (mobile, address, supplier_id))
            else:
                cur.execute("""
                    INSERT INTO suppliers (name, mobile, address, created_at)
                    VALUES (%s,%s,%s,NOW())
                """, (name, mobile, address))
                supplier_id = cur.lastrowid

            total_bill_amount = 0

            if products_json:
                products = json.loads(products_json)

                for p in products:
                    item_name = (p.get("item_name") or "").strip()
                    unit = p.get("unit") or "Pcs"
                    qty = float(p.get("qty") or 0)
                    purchase_price = float(p.get("purchase_price") or 0)
                    selling_price = float(p.get("selling_price") or 0)
                    stock_type = (p.get("stock_type") or "REGULAR").strip().upper()
                    gst_rate = float(p.get("gst_rate") or 0)

                    if stock_type not in ["REGULAR", "GST"]:
                        stock_type = "REGULAR"

                    taxable_value = qty * purchase_price

                    tax_type = (p.get("tax_type") or "INTRA").strip().upper()

                    if stock_type == "GST" and gst_rate > 0:
                        tax_data = calculate_gst_split(taxable_value, gst_rate, tax_type)
                        gst_amount = tax_data["gst_amount"]
                        cgst = tax_data["cgst"]
                        sgst = tax_data["sgst"]
                        igst = tax_data["igst"]
                        total_amount = taxable_value + gst_amount
                    else:
                        gst_rate = 0
                        cgst = 0
                        sgst = 0
                        igst = 0
                        total_amount = taxable_value
                        tax_type = "INTRA"
                    amount = float(p.get("amount") or total_amount)

                    if not item_name or qty <= 0:
                        continue

                    total_bill_amount += amount

                    cur.execute("""
                        SELECT id
                        FROM inventory
                        WHERE LOWER(product_name)=%s AND stock_type=%s
                        LIMIT 1
                    """, (item_name.lower(), stock_type))
                    existing_item = cur.fetchone()

                    if existing_item:
                        cur.execute("""
                            UPDATE inventory
                            SET stock_quantity = stock_quantity + %s,
                                purchase_price=%s,
                                selling_price=%s,
                                supplier_id=%s,
                                unit=%s,
                                stock_type=%s,
                                gst_rate=%s,
                                taxable_value=%s,
                                cgst=%s,
                                sgst=%s,
                                total_amount=%s,
                                purchase_date=%s,
                                rate_updated_at=NOW()
                            WHERE id=%s
                        """, (
                            qty,
                            purchase_price,
                            selling_price,
                            supplier_id,
                            unit,
                            stock_type,
                            gst_rate,
                            taxable_value,
                            cgst,
                            sgst,
                            amount,
                            bill_date,
                            existing_item["id"]
                        ))
                    else:
                        cur.execute("""
                            INSERT INTO inventory
                            (product_name, unit, purchase_price, selling_price,
                             stock_quantity, min_stock_level, stock_type,
                             supplier_id, gst_rate, taxable_value, cgst, sgst,
                             total_amount, purchase_date, rate_updated_at)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
                        """, (
                            item_name,
                            unit,
                            purchase_price,
                            selling_price,
                            qty,
                            0,
                            stock_type,
                            supplier_id,
                            gst_rate,
                            taxable_value,
                            cgst,
                            sgst,
                            amount,
                            bill_date
                        ))

                if total_bill_amount > 0:
                    cur.execute("""
                        INSERT INTO supplier_ledger
                        (supplier_id, bill_no, credit_amount, debit_amount, balance_amount, transaction_date)
                        VALUES (%s,%s,%s,%s,%s,%s)
                    """, (
                        supplier_id,
                        bill_no if bill_no else f"PURCHASE-{supplier_id}",
                        total_bill_amount,
                        0,
                        total_bill_amount,
                        bill_date
                    ))

            mysql.connection.commit()
            flash("Supplier and products saved successfully!", "success")
            return redirect(url_for("suppliers_page"))

        except Exception as e:
            mysql.connection.rollback()
            print("ADD SUPPLIER ERROR:", e)
            flash("Error saving supplier", "danger")

    cur.close()
    return render_template("add_supplier.html")

def calculate_gst_split(taxable_value, gst_rate, tax_type="INTRA"):
    taxable_value = float(taxable_value or 0)
    gst_rate = float(gst_rate or 0)

    gst_amount = (taxable_value * gst_rate) / 100

    if tax_type == "INTER":
        return {
            "gst_amount": gst_amount,
            "cgst": 0,
            "sgst": 0,
            "igst": gst_amount
        }
    else:
        return {
            "gst_amount": gst_amount,
            "cgst": gst_amount / 2,
            "sgst": gst_amount / 2,
            "igst": 0
        }

@app.route("/pay-supplier/<int:supplier_id>", methods=["GET","POST"])
@login_required
def pay_supplier(supplier_id):

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # supplier info
    cur.execute("SELECT * FROM suppliers WHERE id=%s",(supplier_id,))
    supplier = cur.fetchone()


    # ===============================
    # TOTAL PURCHASE FROM PURCHASE BILLS
    # ===============================
    cur.execute("""
    SELECT IFNULL(SUM(purchase_price * stock_quantity),0) AS total_purchase
    FROM inventory
    WHERE supplier_id=%s
    """, (supplier_id,))
    row = cur.fetchone()
    total_purchase = row["total_purchase"] if row else 0

    # ===============================
    # PAID FROM SUPPLIER PAYMENTS
    # ===============================
    cur.execute("""
    SELECT COALESCE(SUM(amount),0) AS paid_extra
    FROM supplier_payments
    WHERE supplier_id=%s
    """, (supplier_id,))
    paid_extra = cur.fetchone()["paid_extra"]

    total_paid = float(paid_extra)

    balance = float(total_purchase) - float(total_paid)

    if request.method == "POST":

        amount = float(request.form.get("amount"))
        method = request.form.get("payment_method")
        paid_by = request.form.get("paid_by")
        staff = request.form.get("staff")
        notes = request.form.get("notes")

        # prevent over payment
        if amount > balance:
            return redirect(url_for("pay_supplier", supplier_id=supplier_id))

        # save payment
        cur.execute("""
        INSERT INTO supplier_payments
        (supplier_id, amount, payment_method, paid_by, staff_name, notes, payment_date)
        VALUES (%s,%s,%s,%s,%s,%s,NOW())
        """,(supplier_id,amount,method,paid_by,staff,notes))


        mysql.connection.commit()


        return redirect(url_for("suppliers"))

    cur.close()

    return render_template(
        "pay_supplier.html",
        supplier=supplier,
        balance=balance,
        total_purchase=total_purchase,
        total_paid=total_paid
    )

@app.route("/supplier-details/<int:supplier_id>")
@login_required
def supplier_details(supplier_id):

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cur.execute("SELECT * FROM suppliers WHERE id=%s",(supplier_id,))
    supplier = cur.fetchone()

    cur.execute("""
        SELECT *
        FROM supplier_ledger
        WHERE supplier_id=%s
        ORDER BY transaction_date DESC
    """,(supplier_id,))
    entries = cur.fetchall()

    cur.close()

    return render_template(
        "supplier_details.html",
        supplier=supplier,
        entries=entries
    )
@app.route("/quick-pay-supplier", methods=["POST"])
@login_required
def quick_pay_supplier():

    try:

        data = request.get_json()

        supplier_id = data.get("supplier_id")
        supplier_name = data.get("supplier_name")
        amount = float(data.get("amount") or 0)
        method = data.get("method") or "Cash"

        cur = mysql.connection.cursor()

        # create supplier if missing
        if not supplier_id:

            cur.execute(
                "SELECT id FROM suppliers WHERE name=%s",
                (supplier_name,)
            )
            row = cur.fetchone()

            if row:
                supplier_id = row[0]
            else:
                cur.execute(
                    "INSERT INTO suppliers (name,created_at) VALUES (%s,NOW())",
                    (supplier_name,)
                )
                supplier_id = cur.lastrowid

        # insert payment
        cur.execute("""
        INSERT INTO supplier_payments
        (supplier_id, amount, payment_method, payment_date, created_at)
        VALUES (%s,%s,%s,NOW(),NOW())
        """,(supplier_id,amount,method))

        mysql.connection.commit()
        cur.close()

        return jsonify({"success":True})

    except Exception as e:

        print("SUPPLIER PAYMENT ERROR:", e)

        return jsonify({
            "success":False,
            "error":str(e)
        })

@app.route("/api/supplier-balance/<int:supplier_id>")
@login_required
def supplier_balance(supplier_id):

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    try:

        # total purchase from inventory
        cur.execute("""
        SELECT IFNULL(SUM(purchase_price * stock_quantity),0) AS total_purchase
        FROM purchase_bills
        WHERE supplier_id=%s
        """,(supplier_id,))

        purchase = cur.fetchone()["total_purchase"] or 0


        # total already paid
        cur.execute("""
        SELECT IFNULL(SUM(amount),0) AS total_paid
        FROM supplier_payments
        WHERE supplier_id=%s
        """,(supplier_id,))

        paid = cur.fetchone()["total_paid"] or 0


        balance = float(purchase) - float(paid)

        return jsonify({
            "total_purchase": purchase,
            "total_paid": paid,
            "balance": balance
        })

    except Exception as e:

        print("Supplier balance error:", e)

        return jsonify({
            "total_purchase":0,
            "total_paid":0,
            "balance":0
        })

    finally:
        cur.close()

@app.route("/purchase", methods=["GET","POST"])
@login_required
def purchase():

    cur = mysql.connection.cursor()

    if request.method == "POST":

        supplier_id = request.form["supplier_id"]

        cur.execute("""
        INSERT INTO purchase_bills (supplier_id,bill_date)
        VALUES (%s,NOW())
        """,(supplier_id,))

        purchase_id = cur.lastrowid

        pids = request.form.getlist("product_id[]")
        qtys = request.form.getlist("qty[]")
        rates = request.form.getlist("rate[]")

        for i in range(len(pids)):

            pid = pids[i]
            q = float(qtys[i])
            r = float(rates[i])

            total = q*r

            cur.execute("""
            INSERT INTO purchase_items
            (purchase_id,product_id,qty,rate,total)
            VALUES (%s,%s,%s,%s,%s)
            """,(purchase_id,pid,q,r,total))

            cur.execute("""
            UPDATE inventory
            SET stock_quantity = stock_quantity + %s
            WHERE id=%s
            """,(q,pid))

        mysql.connection.commit()

    return render_template("purchase.html")

@app.route("/api/quick-add-supplier", methods=["POST"])
def quick_add_supplier():
    name = request.json.get('name')
    cur = mysql.connection.cursor()
    cur.execute("INSERT INTO suppliers (name) VALUES (%s)", (name,))
    new_id = cur.lastrowid
    mysql.connection.commit()
    cur.close()
    return jsonify({"id": new_id, "name": name})

@app.route("/add-supplier-fast", methods=["POST"])
def add_supplier_fast():
    data = request.get_json()
    name = data.get('name')
    cur = mysql.connection.cursor()
    cur.execute("INSERT INTO suppliers (name) VALUES (%s)", (name,))
    mysql.connection.commit()
    new_id = cur.lastrowid
    cur.close()
    return jsonify({"id": new_id, "name": name})


@app.route("/credit-ledger")
@login_required
def credit_ledger():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cur.execute("""
        SELECT 
            customer_name,
            customer_mobile,
            SUM(total_udhari) AS total_udhari,
            SUM(pending_bills) AS pending_bills,
            MAX(last_transaction) AS last_transaction,
            GROUP_CONCAT(DISTINCT source) AS sources
        FROM
        (
            SELECT 
                customer_name,
                customer_mobile,
                SUM(GREATEST(balance_amount,0)) AS total_udhari,
                COUNT(id) AS pending_bills,
                MAX(created_at) AS last_transaction,
                'BILL' AS source
            FROM bills
            WHERE balance_amount > 0
              AND status IN ('DONE','BAKI')
            GROUP BY customer_name, customer_mobile

            UNION ALL

            SELECT 
                customer_name,
                customer_mobile,
                SUM(amount) AS total_udhari,
                COUNT(id) AS pending_bills,
                MAX(created_at) AS last_transaction,
                'CHILLAR' AS source
            FROM chillar_entries
            WHERE type = 'BAKI'
            AND status = 'BAKI'
            GROUP BY customer_name, customer_mobile
        ) AS combined
        GROUP BY customer_name, customer_mobile
        ORDER BY total_udhari DESC
    """)

    ledger_summary = cur.fetchall()

    grand_total = sum(float(r["total_udhari"]) for r in ledger_summary)

    cur.close()

    return render_template(
        "credit_ledger.html",
        ledger=ledger_summary,
        grand_total=grand_total
    )

@app.route("/receive-credit-ledger", methods=["POST"])
@login_required
def receive_credit_ledger():

    name = request.form.get("customer_name")
    mobile = request.form.get("customer_mobile")
    amount_paid = float(request.form.get("amount_paid") or 0)
    payment_mode = request.form.get("payment_mode")
    upi_account = request.form.get("upi_account")
    staff_id = request.form.get("staff_id") if upi_account == "Staff" else None
    cur = mysql.connection.cursor()

    remaining = amount_paid

    # =========================
    # CLEAR BILL UDHARI FIRST
    # =========================
    cur.execute("""
        SELECT id, balance_amount
        FROM bills
        WHERE customer_name=%s
          AND customer_mobile=%s
          AND balance_amount>0
          AND status!='ESTIMATE'
        ORDER BY created_at ASC
    """, (name, mobile))

    bills = cur.fetchall()

    for row in bills:
        if remaining <= 0:
            break

        bid = row[0]
        bal = float(row[1])

        if remaining >= bal:
            cur.execute("""
            UPDATE bills
            SET paid_amount = paid_amount + %s,
                balance_amount = 0,
                payment_mode=%s,
                upi_account=%s,
                upi_staff_id=%s
            WHERE id=%s
            """, (bal, payment_mode, upi_account, staff_id, bid))
            remaining -= bal
        else:
            cur.execute("""
            UPDATE bills
            SET paid_amount = paid_amount + %s,
                balance_amount = balance_amount - %s,
                payment_mode=%s,
                upi_account=%s,
                upi_staff_id=%s
            WHERE id=%s
            """, (remaining, remaining, payment_mode, upi_account, staff_id, bid))
            remaining = 0

    # =========================
    # CLEAR CHILLAR BAKI
    # =========================
    if remaining > 0:
        cur.execute("""
            SELECT id, amount
            FROM chillar_entries
            WHERE customer_name=%s
              AND customer_mobile=%s
              AND type='BAKI'
              AND status='BAKI'
            ORDER BY created_at ASC
        """, (name, mobile))

        baki_rows = cur.fetchall()

        for row in baki_rows:
            if remaining <= 0:
                break

            cid = row[0]
            amt = float(row[1])

            if remaining >= amt:
                cur.execute("""
                UPDATE chillar_entries
                SET status='PAID',
                    received_at=NOW(),
                    payment_mode=%s,
                    upi_account=%s
                WHERE id=%s
                """, (payment_mode, upi_account, cid))
                remaining -= amt
            else:
                new_amt = amt - remaining
                cur.execute("""
                INSERT INTO chillar_entries
                (customer_name, customer_mobile,
                 amount, payment_mode,
                 upi_account,
                 type, status, created_at)
                VALUES (%s,%s,%s,%s,%s,'BIKRI','PAID',NOW())
                """, (name, mobile, remaining, payment_mode, upi_account))
                remaining = 0

    mysql.connection.commit()
    cur.close()


    return redirect(url_for("credit_ledger"))

@app.route("/customer-ledger")
@login_required
def customer_ledger():

    name = request.args.get("name")

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # BILLS
    cur.execute("""
        SELECT 
            created_at,
            'BILL' AS source,
            id AS ref_id,
            total_amount AS debit,
            0 AS credit,
            total_amount AS balance  
        FROM bills
        WHERE customer_name=%s
          AND status!='ESTIMATE'
    """, (name,))
    bills = cur.fetchall()

    # CHILLAR
    cur.execute("""
        SELECT 
            created_at,
            'CHILLAR' AS source,
            id AS ref_id,
            amount AS debit,
            0 AS credit,
            amount AS balance
        FROM chillar_entries
        WHERE customer_name=%s
          AND type='BAKI'
    """, (name,))
    chillar = cur.fetchall()

    # ADVANCE
    cur.execute("""
        SELECT
            ab.created_at,
            'ADVANCE' AS source,
            ab.id AS ref_id,
            0 AS debit,
            ab.advance_amount AS credit,
            0 AS balance
        FROM advance_bookings ab
        JOIN customers c ON ab.customer_id = c.id
        WHERE c.name=%s
    """, (name,))
    adv = cur.fetchall()

    ledger = list(bills) + list(chillar) + list(adv)
    ledger = sorted(ledger, key=lambda x: x["created_at"])

    running = 0
    for row in ledger:
        debit = float(row.get("debit") or 0)
        credit = float(row.get("credit") or 0)
        running += debit - credit
        row["running_balance"] = running

    ledger = list(reversed(ledger))

    cur.close()

    return render_template(
        "customer_ledger.html",
        ledger=ledger,
        customer=name
    )

@app.route("/customers", methods=["GET", "POST"])
@login_required
def customers():

    # =========================
    # SEARCH HANDLING
    # =========================

    search = ""

    if request.method == "POST":
        search = request.form.get("q", "").strip()

    page = int(request.args.get("page", 1))
    per_page = 12
    offset = (page - 1) * per_page

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # =========================
    # UNIQUE CUSTOMERS FROM ALL TABLES
    # =========================

    base_union = """
        SELECT customer_name AS name, customer_mobile AS phone FROM bills
        UNION
        SELECT customer_name AS name, customer_mobile AS phone FROM chillar_entries
        UNION
        SELECT name, phone FROM customers
    """

    params = []

    if search:
        base_query = f"""
            SELECT * FROM ({base_union}) AS combined
            WHERE name LIKE %s OR phone LIKE %s
        """
        params = [f"%{search}%", f"%{search}%"]
    else:
        base_query = f"SELECT * FROM ({base_union}) AS combined"

    # =========================
    # COUNT TOTAL CUSTOMERS
    # =========================

    cur.execute(f"""
        SELECT COUNT(*) AS total
        FROM ({base_query}) AS count_table
    """, params)

    total_customers = cur.fetchone()["total"]
    total_pages = (total_customers // per_page) + (1 if total_customers % per_page else 0)

    # =========================
    # FETCH PAGE DATA
    # =========================

    cur.execute(f"""
        SELECT *
        FROM ({base_query}) AS final
        ORDER BY name
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])

    customers = cur.fetchall()

    # =========================
    # ATTACH SALES / UDHARI / ADVANCE
    # =========================

    for c in customers:

        # Bills
        cur.execute("""
            SELECT 
                IFNULL(SUM(total_amount),0) AS total_sales,
                IFNULL(SUM(balance_amount),0) AS total_udhari,
                MAX(created_at) AS last_bill_date
            FROM bills
            WHERE customer_name=%s
            AND customer_mobile=%s
            AND status IN ('DONE','BAKI')
        """, (c["name"], c["phone"]))

        bill_data = cur.fetchone()

        c["total_sales"] = float(bill_data["total_sales"])
        c["total_udhari"] = float(bill_data["total_udhari"])
        c["last_date"] = bill_data["last_bill_date"]

        # Advance
        cur.execute("""
            SELECT IFNULL(SUM(ab.advance_amount),0) AS total_advance
            FROM advance_bookings ab
            JOIN customers cu ON ab.customer_id = cu.id
            WHERE cu.name=%s
            AND cu.phone=%s
        """, (c["name"], c["phone"]))

        adv_data = cur.fetchone()
        c["total_advance"] = float(adv_data["total_advance"])

    # =========================
    # SUMMARY
    # =========================

    cur.execute("""
        SELECT 
            IFNULL(SUM(balance_amount),0) AS total_udhari,
            IFNULL(SUM(total_amount),0) AS total_sales
        FROM bills
        WHERE status IN ('DONE','BAKI')
    """)

    summary = cur.fetchone()

    total_udhari = float(summary["total_udhari"])
    total_sales = float(summary["total_sales"])

    cur.close()

    return render_template(
        "customers.html",
        customers=customers,
        total_customers=total_customers,
        total_udhari=total_udhari,
        total_sales=total_sales,
        page=page,
        total_pages=total_pages
    )


@app.route("/receive-payment/<int:bill_id>", methods=["GET","POST"])
@login_required
def receive_payment(bill_id):
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cur.execute("""
        SELECT id, customer_name, customer_mobile, balance_amount
        FROM bills
        WHERE id=%s
    """, (bill_id,))
    bill = cur.fetchone()

    if not bill:

        return redirect(url_for("pending_bills"))

    if request.method == "POST":
        amount_paid = float(request.form.get("amount_paid") or 0)
        remaining = amount_paid
        customer_name = bill["customer_name"]

        cur.execute("""
            SELECT id, balance_amount
            FROM bills
            WHERE customer_name=%s
            AND balance_amount>0
            AND status!='ESTIMATE'
            ORDER BY created_at ASC
        """, (customer_name,))
        pending = cur.fetchall()

        for row in pending:
            if remaining <= 0:
                break

            bal = float(row["balance_amount"])
            bid = row["id"]

            if remaining >= bal:
                cur.execute("""
                    UPDATE bills
                    SET balance_amount=0,
                        paid_amount=paid_amount+%s
                    WHERE id=%s
                """, (bal, bid))
                remaining -= bal
            else:
                cur.execute("""
                    UPDATE bills
                    SET balance_amount=balance_amount-%s,
                        paid_amount=paid_amount+%s
                    WHERE id=%s
                """, (remaining, remaining, bid))
                remaining = 0

        mysql.connection.commit()
        cur.close()


        return redirect(url_for("pending_bills"))

    cur.close()
    return render_template(
        "receive_payment_form.html",
        bill_id=bill["id"],   # ⭐ ADD THIS
        customer_name=bill["customer_name"],
        total_due=bill["balance_amount"]
    )

from flask import request, render_template, send_file
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

@app.route("/gst-report")
@login_required
def gst_report():
    filter_type = request.args.get("type", "month")
    month = request.args.get("month", "").strip()
    date_filter = request.args.get("date", "").strip()
    year = request.args.get("year", "").strip()
    week = request.args.get("week", "").strip()
    report_type = request.args.get("report", "sale").strip().lower()

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    params = []

    rows = []
    grand_total = 0
    taxable_total = 0
    gst_total = 0
    cgst_total = 0
    sgst_total = 0
    igst_total = 0

    if report_type == "sale":
        query = """
            SELECT
                id AS bill_id,
                customer_name,
                buyer_gstin,
                total_amount,
                taxable_value,
                gst_amount,
                cgst_amount,
                sgst_amount,
                igst_amount,
                tax_type,
                created_at AS bill_date
            FROM bills
            WHERE status != 'ESTIMATE'
              AND bill_type = 'GST'
        """

        if filter_type == "month" and month:
            query += " AND DATE_FORMAT(created_at, '%%Y-%%m') = %s"
            params.append(month)
        elif filter_type == "date" and date_filter:
            query += " AND DATE(created_at) = %s"
            params.append(date_filter)
        elif filter_type == "year" and year:
            query += " AND YEAR(created_at) = %s"
            params.append(year)
        elif filter_type == "week" and week and year:
            query += " AND YEAR(created_at) = %s AND WEEK(created_at, 1) = %s"
            params.extend([year, week])

        query += " ORDER BY created_at DESC"
        cur.execute(query, params)
        rows = cur.fetchall()

        for r in rows:
            taxable = float(r.get("taxable_value") or 0)
            cgst = float(r.get("cgst_amount") or 0)
            sgst = float(r.get("sgst_amount") or 0)
            igst = float(r.get("igst_amount") or 0)
            gst = float(r.get("gst_amount") or (cgst + sgst + igst))
            total = float(r.get("total_amount") or 0)

            r["taxable"] = taxable
            r["cgst"] = cgst
            r["sgst"] = sgst
            r["igst"] = igst
            r["gst_total"] = gst
            r["total"] = total
            r["gst_percent"] = round((gst / taxable) * 100, 2) if taxable > 0 else 0
            r["cgst_percent"] = round((cgst / taxable) * 100, 2) if taxable > 0 else 0
            r["sgst_percent"] = round((sgst / taxable) * 100, 2) if taxable > 0 else 0
            r["igst_percent"] = round((igst / taxable) * 100, 2) if taxable > 0 else 0

            taxable_total += taxable
            cgst_total += cgst
            sgst_total += sgst
            igst_total += igst
            gst_total += gst
            grand_total += total

    else:
        query = """
            SELECT
                i.id,
                i.product_name,
                s.name AS supplier_name,
                i.stock_quantity,
                i.purchase_price,
                i.taxable_value,
                i.cgst,
                i.sgst,
                i.igst,
                i.total_amount,
                i.tax_type,
                i.purchase_date AS bill_date
            FROM inventory i
            LEFT JOIN suppliers s ON i.supplier_id = s.id
            WHERE i.stock_type = 'GST'
        """

        if filter_type == "month" and month:
            query += " AND DATE_FORMAT(i.purchase_date, '%%Y-%%m') = %s"
            params.append(month)
        elif filter_type == "date" and date_filter:
            query += " AND DATE(i.purchase_date) = %s"
            params.append(date_filter)
        elif filter_type == "year" and year:
            query += " AND YEAR(i.purchase_date) = %s"
            params.append(year)
        elif filter_type == "week" and week and year:
            query += " AND YEAR(i.purchase_date) = %s AND WEEK(i.purchase_date, 1) = %s"
            params.extend([year, week])

        query += " ORDER BY i.purchase_date DESC"
        cur.execute(query, params)
        rows = cur.fetchall()

        for r in rows:
            taxable = float(r.get("taxable_value") or 0)
            cgst = float(r.get("cgst") or 0)
            sgst = float(r.get("sgst") or 0)
            igst = float(r.get("igst") or 0)
            gst = cgst + sgst + igst
            total = float(r.get("total_amount") or 0)

            r["taxable"] = taxable
            r["cgst"] = cgst
            r["sgst"] = sgst
            r["igst"] = igst
            r["gst_total"] = gst
            r["total"] = total
            r["gst_percent"] = round((gst / taxable) * 100, 2) if taxable > 0 else 0
            r["cgst_percent"] = round((cgst / taxable) * 100, 2) if taxable > 0 else 0
            r["sgst_percent"] = round((sgst / taxable) * 100, 2) if taxable > 0 else 0
            r["igst_percent"] = round((igst / taxable) * 100, 2) if taxable > 0 else 0

            taxable_total += taxable
            cgst_total += cgst
            sgst_total += sgst
            igst_total += igst
            gst_total += gst
            grand_total += total

    cur.close()

    return render_template(
        "gst_report.html",
        rows=rows,
        report_type=report_type,
        filter_type=filter_type,
        month=month,
        date_filter=date_filter,
        year=year,
        week=week,
        taxable=taxable_total,
        tax=gst_total,
        cgst_total=cgst_total,
        sgst_total=sgst_total,
        igst_total=igst_total,
        grand=grand_total
    )

@app.route("/gst-report-excel")
@login_required
def gst_report_excel():
    filter_type = request.args.get("type", "month")
    month = request.args.get("month", "").strip()
    date_filter = request.args.get("date", "").strip()
    year = request.args.get("year", "").strip()
    week = request.args.get("week", "").strip()
    report_type = request.args.get("report", "sale").strip().lower()

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    params = []
    wb = Workbook()
    ws = wb.active
    ws.title = "GST Report"

    if report_type == "purchase":
        query = """
            SELECT
                i.purchase_date,
                i.id,
                i.product_name,
                s.name AS supplier_name,
                i.stock_quantity,
                i.purchase_price,
                i.taxable_value,
                i.cgst,
                i.sgst,
                i.igst,
                i.total_amount,
                i.tax_type
            FROM inventory i
            LEFT JOIN suppliers s ON i.supplier_id = s.id
            WHERE i.stock_type = 'GST'
        """

        if filter_type == "month" and month:
            query += " AND DATE_FORMAT(i.purchase_date, '%%Y-%%m') = %s"
            params.append(month)
        elif filter_type == "date" and date_filter:
            query += " AND DATE(i.purchase_date) = %s"
            params.append(date_filter)
        elif filter_type == "year" and year:
            query += " AND YEAR(i.purchase_date) = %s"
            params.append(year)
        elif filter_type == "week" and week and year:
            query += " AND YEAR(i.purchase_date) = %s AND WEEK(i.purchase_date, 1) = %s"
            params.extend([year, week])

        query += " ORDER BY i.purchase_date DESC"
        cur.execute(query, params)
        rows = cur.fetchall()

        headers = [
            "DATE", "ID", "SUPPLIER", "PRODUCT", "QTY", "PUR RATE",
            "TAX TYPE", "GST %", "TAXABLE", "CGST", "SGST", "IGST",
            "GST TOTAL", "TOTAL"
        ]
        ws.append(headers)

        for r in rows:
            taxable = round(float(r.get("taxable_value") or 0), 2)
            cgst = round(float(r.get("cgst") or 0), 2)
            sgst = round(float(r.get("sgst") or 0), 2)
            igst = round(float(r.get("igst") or 0), 2)
            gst_total = round(cgst + sgst + igst, 2)
            gst_percent = round((gst_total / taxable) * 100, 2) if taxable > 0 else 0

            ws.append([
                r["purchase_date"].strftime("%d-%m-%Y") if r.get("purchase_date") else "",
                r.get("id", ""),
                r.get("supplier_name", ""),
                r.get("product_name", ""),
                float(r.get("stock_quantity") or 0),
                round(float(r.get("purchase_price") or 0), 2),
                r.get("tax_type", ""),
                gst_percent,
                taxable,
                cgst,
                sgst,
                igst,
                gst_total,
                round(float(r.get("total_amount") or 0), 2)
            ])

    else:
        query = """
            SELECT
                id,
                customer_name,
                buyer_gstin,
                taxable_value,
                cgst_amount,
                sgst_amount,
                igst_amount,
                gst_amount,
                total_amount,
                tax_type,
                created_at
            FROM bills
            WHERE bill_type = 'GST'
              AND status != 'ESTIMATE'
        """

        if filter_type == "month" and month:
            query += " AND DATE_FORMAT(created_at, '%%Y-%%m') = %s"
            params.append(month)
        elif filter_type == "date" and date_filter:
            query += " AND DATE(created_at) = %s"
            params.append(date_filter)
        elif filter_type == "year" and year:
            query += " AND YEAR(created_at) = %s"
            params.append(year)
        elif filter_type == "week" and week and year:
            query += " AND YEAR(created_at) = %s AND WEEK(created_at, 1) = %s"
            params.extend([year, week])

        query += " ORDER BY created_at DESC"
        cur.execute(query, params)
        rows = cur.fetchall()

        headers = [
            "DATE", "BILL NO", "CUSTOMER", "GSTIN", "TAX TYPE", "GST %",
            "TAXABLE", "CGST", "SGST", "IGST", "GST TOTAL", "TOTAL"
        ]
        ws.append(headers)

        for r in rows:
            taxable = round(float(r.get("taxable_value") or 0), 2)
            cgst = round(float(r.get("cgst_amount") or 0), 2)
            sgst = round(float(r.get("sgst_amount") or 0), 2)
            igst = round(float(r.get("igst_amount") or 0), 2)
            gst_total = round(float(r.get("gst_amount") or (cgst + sgst + igst)), 2)
            gst_percent = round((gst_total / taxable) * 100, 2) if taxable > 0 else 0

            ws.append([
                r["created_at"].strftime("%d-%m-%Y") if r.get("created_at") else "",
                r.get("id", ""),
                r.get("customer_name", ""),
                r.get("buyer_gstin", ""),
                r.get("tax_type", ""),
                gst_percent,
                taxable,
                cgst,
                sgst,
                igst,
                gst_total,
                round(float(r.get("total_amount") or 0), 2)
            ])

    cur.close()

    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max_len + 3

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"gst_{report_type}_report.xlsx"
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/gst-register-excel")
@login_required
def gst_register_excel():

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cur.execute("""
        SELECT 
            b.id,
            b.customer_name,
            b.buyer_gstin,
            b.taxable_value,
            b.cgst_amount,
            b.sgst_amount,
            b.igst_amount,
            b.gst_amount,
            b.total_amount,
            b.tax_type,
            b.created_at
        FROM bills b
        WHERE b.bill_type='GST'
          AND b.status!='ESTIMATE'
        ORDER BY b.created_at DESC
    """)
    sales = cur.fetchall()

    cur.execute("""
        SELECT
            i.product_name,
            i.purchase_price,
            i.stock_quantity,
            s.name AS supplier_name,
            i.taxable_value,
            i.cgst,
            i.sgst,
            i.igst,
            i.total_amount,
            i.tax_type,
            i.purchase_date
        FROM inventory i
        LEFT JOIN suppliers s ON i.supplier_id=s.id
        WHERE i.stock_type='GST'
        ORDER BY i.purchase_date DESC
    """)
    purchase = cur.fetchall()

    cur.close()

    wb = Workbook()

    # PURCHASE SHEET
    ws = wb.active
    ws.title = "GST Purchase"
    ws.append([
        "DATE",
        "SUPPLIER",
        "PRODUCT",
        "QTY",
        "PUR RATE",
        "TAXABLE",
        "CGST",
        "SGST",
        "IGST",
        "TAX TYPE",
        "TOTAL"
    ])

    for p in purchase:
        ws.append([
            p["purchase_date"],
            p["supplier_name"],
            p["product_name"],
            float(p["stock_quantity"] or 0),
            float(p["purchase_price"] or 0),
            round(float(p["taxable_value"] or 0), 2),
            round(float(p["cgst"] or 0), 2),
            round(float(p["sgst"] or 0), 2),
            round(float(p["igst"] or 0), 2),
            p.get("tax_type", ""),
            round(float(p["total_amount"] or 0), 2)
        ])

    # SALES SHEET
    ws2 = wb.create_sheet("GST Sales")
    ws2.append([
        "DATE",
        "BILL NO",
        "CUSTOMER",
        "GSTIN",
        "TAXABLE",
        "CGST",
        "SGST",
        "IGST",
        "GST TOTAL",
        "TAX TYPE",
        "TOTAL"
    ])

    for s in sales:
        ws2.append([
            s["created_at"],
            s["id"],
            s["customer_name"],
            s.get("buyer_gstin", ""),
            round(float(s["taxable_value"] or 0), 2),
            round(float(s["cgst_amount"] or 0), 2),
            round(float(s["sgst_amount"] or 0), 2),
            round(float(s["igst_amount"] or 0), 2),
            round(float(s["gst_amount"] or 0), 2),
            s.get("tax_type", ""),
            round(float(s["total_amount"] or 0), 2)
        ])

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)

    return send_file(tmp.name, as_attachment=True, download_name="gst_register.xlsx")

@app.route("/billing")
@login_required
def billing():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    edit_id = request.args.get("edit")
    bill = None
    items = []
    cur.execute("SELECT * FROM hamals")
    hamals = cur.fetchall()

    if edit_id:
        # Get bill details
        cur.execute("SELECT * FROM bills WHERE id=%s", (edit_id,))
        bill = cur.fetchone()

        # Get bill items
        cur.execute("SELECT * FROM bill_items WHERE bill_id=%s", (edit_id,))
        items = cur.fetchall()

    # ✅ OWNER UPI
    cur.execute("SELECT upi_id FROM users WHERE role='owner' LIMIT 1")
    row = cur.fetchone()
    owner_upi = row["upi_id"] if row else ""

    # ✅ STAFF LIST WITH UPI
    cur.execute("SELECT id,name,upi_id FROM users WHERE role='staff'")
    staff = cur.fetchall()

    cur.execute("""
    SELECT id FROM daily_closing
    WHERE closing_date = CURDATE()
    """)

    if cur.fetchone():

        return redirect(url_for("dashboard"))

    cur.close()

    return render_template(
        "billing.html",
        hamals=hamals,
        staff=staff,
        owner_upi=owner_upi,
        role=session.get("role"),
        edit_id=edit_id,
        bill=bill,
        items=items
    )

@app.route("/print-bill/<int:bill_id>")
@login_required
def print_bill(bill_id):

    from num2words import num2words

    def number_to_words(n):
        return num2words(n, lang='en_IN').title()

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # ================= GET BILL =================
    cur.execute("SELECT * FROM bills WHERE id=%s",(bill_id,))
    bill = cur.fetchone()

    if not bill:
        cur.close()
        return "Bill not found", 404


    # ================= GET BILL ITEMS =================
    cur.execute("""
    SELECT 
        bi.quantity,
        bi.rate,
        bi.unit,
        COALESCE(i.product_name, bi.product_name) AS product_name,
        bi.hsn_code
    FROM bill_items bi
    LEFT JOIN inventory i ON bi.product_id = i.id
    WHERE bi.bill_id=%s
    """, (bill_id,))

    items = cur.fetchall()


    # ================= GET HAMALI =================
    cur.execute("""
    SELECT IFNULL(SUM(credit_amount),0) AS hamali_total
    FROM hamal_ledger
    WHERE bill_id=%s
    AND description LIKE 'Bill %% Hamali'
    """, (bill_id,))

    hamali_total = cur.fetchone()["hamali_total"]


    # ================= BHADA =================
    bhada = bill.get("bhada_amount",0)

    cur.close()

    # ================= AMOUNT IN WORDS =================
    amount_words = number_to_words(int(bill["total_amount"]))


    # ================= BILL TYPE =================
    if bill["bill_type"] == "ESTIMATE":
        return render_template(
            "print_estimate.html",
            bill=bill,
            items=items,
            amount_words=amount_words
        )

    if bill["bill_type"] == "GST":
        return render_template(
            "print_gst_bill.html",
            bill=bill,
            items=items,
            amount_words=amount_words
        )

    # ================= NORMAL BILL =================
    return render_template(
        "print_bill.html",
        bill=bill,
        items=items,
        hamali_total=hamali_total,
        bhada=bhada,
        amount_words=amount_words
    )

@app.route("/receipts")
@login_required
def receipts():
    q = request.args.get("q", "").strip()
    f_type = request.args.get("type", "").strip()
    f_status = request.args.get("status", "").strip()
    from_date = request.args.get("from", "").strip()
    to_date = request.args.get("to", "").strip()

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    query = """
        SELECT 
            id,
            customer_name,
            customer_mobile,
            buyer_gstin,
            total_amount,
            paid_amount,
            balance_amount,
            bill_type,
            status,
            taxable_value,
            gst_rate,
            gst_amount,
            cgst_amount,
            sgst_amount,
            igst_amount,
            tax_type,
            created_at
        FROM bills
        WHERE 1=1
    """
    params = []

    if q:
        query += """
            AND (
                customer_name LIKE %s
                OR customer_mobile LIKE %s
                OR CAST(id AS CHAR) LIKE %s
            )
        """
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]

    # BILL TYPE FILTER
    if f_type:
        query += " AND bill_type = %s"
        params.append(f_type)

    # STATUS FILTER
    if f_status == "DONE":
        query += " AND balance_amount <= 0"

    elif f_status == "BAKI":
        query += " AND balance_amount > 0"

    if from_date:
        query += " AND DATE(created_at) >= %s"
        params.append(from_date)

    if to_date:
        query += " AND DATE(created_at) <= %s"
        params.append(to_date)

    query += " ORDER BY id DESC LIMIT 300"

    cur.execute(query, params)
    bills = cur.fetchall()

    cur.execute("""
        SELECT IFNULL(SUM(balance_amount), 0) AS total_outstanding
        FROM bills
        WHERE balance_amount > 0
          AND bill_type != 'ESTIMATE'
    """)
    total_outstanding = cur.fetchone()["total_outstanding"]

    cur.execute("""
        SELECT IFNULL(SUM(paid_amount), 0) AS collected_today
        FROM bills
        WHERE DATE(created_at) = CURDATE()
          AND bill_type != 'ESTIMATE'
    """)
    collected_today = cur.fetchone()["collected_today"]

    cur.close()

    return render_template(
        "receipts.html",
        bills=bills,
        total_outstanding=total_outstanding,
        collected_today=collected_today
    )

@app.route("/delete-bill/<int:bill_id>", methods=["POST"])
@login_required
def delete_bill(bill_id):

    cur = mysql.connection.cursor()

    try:
        cur.execute(
            "SELECT product_id, quantity FROM bill_items WHERE bill_id=%s",
            (bill_id,)
        )
        items = cur.fetchall()

        for pid, qty in items:
            cur.execute(
                "UPDATE inventory SET stock_quantity = stock_quantity + %s WHERE id=%s",
                (qty, pid)
            )

        cur.execute("DELETE FROM bill_items WHERE bill_id=%s", (bill_id,))
        cur.execute("DELETE FROM bills WHERE id=%s", (bill_id,))

        mysql.connection.commit()


        return redirect(url_for("receipts"))

    except Exception as e:
        mysql.connection.rollback()

        return redirect(url_for("receipts"))

    finally:
        cur.close()

@app.route("/delete-all-drafts", methods=["POST"])
@login_required
def delete_all_drafts():
    cur = mysql.connection.cursor()
    try:
        # get all pending bills
        cur.execute("SELECT id FROM bills WHERE status='PENDING'")
        bills = cur.fetchall()

        for (bill_id,) in bills:

            # restore stock
            cur.execute(
                "SELECT product_id, quantity FROM bill_items WHERE bill_id=%s",
                (bill_id,)
            )
            items = cur.fetchall()

            for pid, qty in items:
                cur.execute(
                    "UPDATE inventory SET stock_quantity = stock_quantity + %s WHERE id=%s",
                    (qty, pid)
                )

            # delete items
            cur.execute("DELETE FROM bill_items WHERE bill_id=%s", (bill_id,))

            # delete bill
            cur.execute("DELETE FROM bills WHERE id=%s", (bill_id,))

        mysql.connection.commit()

        return jsonify({"success": True})

    except Exception as e:
        mysql.connection.rollback()
        return jsonify({"success": False, "msg": str(e)})

    finally:
        cur.close()

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from io import BytesIO
from flask import Response

@app.route("/pdf-bill/<int:bill_id>")
@login_required
def pdf_bill(bill_id):

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cur.execute("SELECT * FROM bills WHERE id=%s",(bill_id,))
    bill = cur.fetchone()

    cur.execute("""
        SELECT bi.*, i.product_name
        FROM bill_items bi
        JOIN inventory i ON bi.product_id=i.id
        WHERE bi.bill_id=%s
    """,(bill_id,))
    items = cur.fetchall()

    cur.close()

    if not bill:
        return "Bill not found",404

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer,pagesize=A4)

    width,height = A4
    y = height-40

    # =====================================================
    # WATERMARK FOR ESTIMATE
    # =====================================================
    if bill["bill_type"]=="ESTIMATE":

        pdf.saveState()
        pdf.setFont("Helvetica-Bold",80)
        pdf.setFillGray(0.9,0.3)
        pdf.translate(width/2,height/2)
        pdf.rotate(45)
        pdf.drawCentredString(0,0,"QUOTATION")
        pdf.restoreState()

    # =====================================================
    # HEADER
    # =====================================================
    if bill["bill_type"]=="ESTIMATE":

        pdf.setFont("Helvetica-Bold", 22)
        pdf.drawCentredString(width / 2, y, "QUOTATION")

        y -= 25

        pdf.setFont("Helvetica", 11)

        pdf.drawString(40, y, "Customer:")
        pdf.drawString(110, y, bill["customer_name"] or "")

        pdf.drawRightString(
            width - 40,
            y,
            f"Date: {bill['created_at'].strftime('%d %b %Y')}"
        )

        y -= 20

        pdf.drawString(40, y, f"Quotation No: {bill['id']}")

        y -= 25
        y-=30

        pdf.setFont("Helvetica",11)
        pdf.drawString(40,y,"To,")
        y-=15

        pdf.setFont("Helvetica-Bold",12)
        pdf.drawString(40,y,bill["customer_name"] or "")
        y-=25

        pdf.setFont("Helvetica",10)
        pdf.drawRightString(width-40,y,
            f"Date: {bill['created_at'].strftime('%d %b %Y')}")
        y-=25

    else:

        pdf.setFont("Helvetica-Bold",16)
        pdf.drawCentredString(width/2,y,"SR STEEL & HARDWARE")
        y-=18

        pdf.setFont("Helvetica",10)
        pdf.drawCentredString(width/2,y,"Main Road, Amravati")
        y-=25

        pdf.drawString(40,y,f"Bill No: {bill['id']}")
        pdf.drawRightString(width-40,y,
            f"Date: {bill['created_at'].strftime('%d %b %Y')}")
        y-=15

        pdf.drawString(40,y,f"Customer: {bill['customer_name']}")
        y-=25

    # =====================================================
    # TABLE HEADER
    # =====================================================
    pdf.setFont("Helvetica-Bold", 11)

    pdf.drawString(40, y, "Item")
    pdf.drawRightString(420, y, "Qty")
    pdf.drawRightString(520, y, "Amount")

    y -= 10
    pdf.line(40, y, 550, y)
    y -= 18
    pdf.setFont("Helvetica",10)

    # ITEMS
    for item in items:
        amt=float(item["quantity"])*float(item["rate"])

        pdf.drawString(40, y, str(item["product_name"])[:35])

        pdf.drawRightString(
            420,
            y,
            f'{item["quantity"]} {item["unit"]}'
        )

        pdf.drawRightString(
            520,
            y,
            f"{amt:.2f}"
        )
        y-=15

        if y<80:
            pdf.showPage()
            pdf.setFont("Helvetica",10)
            y=height-40

    y-=10
    pdf.line(300,y,550,y)
    y-=20

    label="Estimated Total" if bill["bill_type"]=="ESTIMATE" else "Total"

    pdf.setFont("Helvetica-Bold",11)
    items_total = float(bill["total_amount"]) - float(bill.get("bhada_amount", 0))
    bhada = float(bill.get("bhada_amount", 0))

    pdf.drawRightString(520, y, f"Items: ₹ {items_total:.2f}")
    y -= 15

    if bhada > 0:
        pdf.drawRightString(520, y, f"Bhada: ₹ {bhada:.2f}")
        y -= 15

    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawRightString(520, y, f"Total: ₹ {float(bill['total_amount']):.2f}")

    if bill["bill_type"]!="ESTIMATE":
        y-=18
        pdf.setFont("Helvetica",10)
        pdf.drawRightString(520,y,
            f"Paid: ₹ {float(bill['paid_amount']):.2f}")
        y-=15
        pdf.drawRightString(520,y,
            f"Balance: ₹ {float(bill['balance_amount']):.2f}")

    pdf.showPage()
    pdf.save()

    buffer.seek(0)

    return Response(
        buffer,
        mimetype="application/pdf",
        headers={
            "Content-Disposition":
            f"inline; filename={'estimate' if bill['bill_type']=='ESTIMATE' else 'bill'}_{bill_id}.pdf"
        }
    )

@app.route("/estimate-history")
@login_required
def estimate_history():

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cur.execute("""
        SELECT 
            b.id,
            b.customer_name,
            b.customer_mobile,
            b.total_amount,
            b.created_at,
            u.name AS staff
        SELECT 
            e.id,
            e.customer_name,
            e.customer_mobile,
            e.total_amount,
            e.created_at,
            u.name AS staff
        FROM estimates e
        LEFT JOIN users u ON e.created_by=u.id
         ORDER BY b.created_at DESC
                LIMIT 200
            """)

    estimates = cur.fetchall()

    cur.close()

    return render_template(
        "estimate_history.html",
        estimates=estimates
    )

@app.route('/rate-update')
@login_required
def rate_update_page():

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # products
    cur.execute("""
        SELECT id, product_name, selling_price, stock_quantity
        FROM inventory
        ORDER BY product_name
    """)
    products = cur.fetchall()

    # owner stats
    staff_stats = []
    if session.get("role") == "owner":
        cur.execute("""
            SELECT u.username,
                   u.is_active,
                   COUNT(r.id) AS edits
            FROM users u
            LEFT JOIN rate_edit_log r
                 ON u.id = r.user_id
            GROUP BY u.id
            ORDER BY edits DESC
        """)
        staff_stats = cur.fetchall()

    # recent price edits
    rate_edits = []
    if session.get("role") == "owner":
        cur.execute("""
            SELECT r.id,
                   u.username,
                   i.product_name,
                   r.old_price,
                   r.new_price,
                   r.changed_at
            FROM rate_edit_log r
            JOIN users u ON r.user_id = u.id
            JOIN inventory i ON r.product_id = i.id
            ORDER BY r.changed_at DESC
            LIMIT 20
        """)
        rows = cur.fetchall()

        for r in rows:
            rate_edits.append({
                "user": r["username"],
                "product": r["product_name"],
                "old": float(r["old_price"]),
                "new": float(r["new_price"]),
                "time": r["changed_at"].strftime("%d %b %H:%M") if r["changed_at"] else ""
            })

    cur.close()

    return render_template(
        "rate_update.html",
        products=products,
        staff_stats=staff_stats,
        rate_edits=rate_edits,
        role=session.get("role")
    )

@app.route('/update-rate', methods=['POST'])
@login_required
def update_rate_action():

    product_id = request.form.get("product_id")
    new_price = request.form.get("new_price")

    if not product_id or not new_price:

        return redirect(url_for("rate_update_page"))

    cur = mysql.connection.cursor()

    # old price
    cur.execute("SELECT selling_price FROM inventory WHERE id=%s", (product_id,))
    old = cur.fetchone()
    old_price = float(old[0]) if old else 0

    # update
    cur.execute("""
        UPDATE inventory
        SET selling_price=%s,
            rate_updated_at=NOW()
        WHERE id=%s
    """, (new_price, product_id))

    # log edit
    cur.execute("""
        INSERT INTO rate_edit_log
        (product_id, user_id, old_price, new_price)
        VALUES (%s,%s,%s,%s)
    """, (product_id, session["user_id"], old_price, new_price))

    mysql.connection.commit()
    cur.close()


    return redirect(url_for("rate_update_page"))


@app.route("/supplier-ledger/<int:supplier_id>")
@login_required
def supplier_ledger(supplier_id):

    page = request.args.get("page", 1, type=int)
    per_page = 50
    offset = (page - 1) * per_page

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # ================= SUPPLIER INFO =================
    cur.execute("SELECT name,mobile FROM suppliers WHERE id=%s",(supplier_id,))
    supplier = cur.fetchone()

    # ================= TOTAL PURCHASE =================
    cur.execute("""
        SELECT IFNULL(SUM(taxable_value),0) AS total_purchase
        FROM inventory
        WHERE supplier_id=%s
    """, (supplier_id,))
    total_purchase = cur.fetchone()["total_purchase"]

    # ================= PAYMENTS =================
    cur.execute("""
        SELECT IFNULL(SUM(amount),0) AS total_paid
        FROM supplier_payments
        WHERE supplier_id=%s
    """,(supplier_id,))
    total_paid = cur.fetchone()["total_paid"]

    balance = float(total_purchase) - float(total_paid)

    # ================= LEDGER ITEMS =================

    cur.execute("""
    SELECT 
    purchase_date as date,
    product_name as bill_no,
    product_name,
    stock_quantity as quantity,
    purchase_price as rate,
    taxable_value as amount,
    0 as paid_amount
    FROM inventory
    WHERE supplier_id=%s

    UNION ALL

    SELECT 
    payment_date as date,
    'Payment' as bill_no,
    '-' as product_name,
    0 as quantity,
    0 as rate,
    0 as amount,
    amount as paid_amount
    FROM supplier_payments
    WHERE supplier_id=%s

    ORDER BY date DESC
    """, (supplier_id, supplier_id))

    rows = cur.fetchall()

    # ================= RUNNING BALANCE =================

    running_balance = 0

    for r in rows:
        purchase = float(r.get("amount", 0) or 0)
        paid = float(r.get("paid_amount", 0) or 0)

        running_balance += purchase
        running_balance -= paid

        r["balance_amount"] = running_balance

    # ================= COUNT =================
    cur.execute("""
        SELECT COUNT(*) as total
        FROM supplier_ledger
        WHERE supplier_id=%s
    """,(supplier_id,))
    total_rows = cur.fetchone()["total"]

    cur.close()

    return render_template(
        "supplier_ledger.html",
        supplier=supplier,
        rows=rows,
        total_purchase=total_purchase,
        total_paid=total_paid,
        balance=balance,
        page=page,
        total_rows=total_rows,
        supplier_id=supplier_id
    )
@app.route("/suppliers")
@login_required
def suppliers():

    page = request.args.get("page", 1, type=int)
    per_page = 25
    offset = (page - 1) * per_page

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # ================= SUPPLIER LIST WITH PURCHASE + PAYMENT =================
    cur.execute("""
        SELECT 
        s.id,
        s.name,
        s.mobile,

        IFNULL(p.total_purchase,0) AS total_purchase,
        IFNULL(sp.total_paid,0) AS total_paid,

        (IFNULL(p.total_purchase,0) - IFNULL(sp.total_paid,0)) AS balance

        FROM suppliers s

        LEFT JOIN (
            SELECT supplier_id,
            SUM(purchase_price * stock_quantity) AS total_purchase
            FROM inventory
            GROUP BY supplier_id
        ) p ON p.supplier_id = s.id

        LEFT JOIN (
            SELECT supplier_id,
            SUM(amount) AS total_paid
            FROM supplier_payments
            GROUP BY supplier_id
        ) sp ON sp.supplier_id = s.id

        ORDER BY s.name
        LIMIT %s OFFSET %s
    """,(per_page,offset))

    suppliers = cur.fetchall()

    # ================= TOTAL SUPPLIER COUNT =================
    cur.execute("SELECT COUNT(*) as total FROM suppliers")
    total_suppliers = cur.fetchone()["total"]

    # ================= TOTAL PURCHASE =================
    cur.execute("""
        SELECT IFNULL(SUM(purchase_price * stock_quantity),0) as total_purchase
        FROM inventory
    """)
    total_purchase = cur.fetchone()["total_purchase"]

    # ================= TOTAL SUPPLIER PAYMENTS =================
    cur.execute("""
        SELECT IFNULL(SUM(amount),0) as total_paid
        FROM supplier_payments
    """)
    total_paid = cur.fetchone()["total_paid"]

    # ================= TOTAL PAYABLE =================
    total_payable = float(total_purchase) - float(total_paid)

    cur.close()

    return render_template(
        "suppliers.html",
        suppliers=suppliers,
        page=page,
        total_suppliers=total_suppliers,
        total_payable=total_payable
    )

@app.route("/api/search-supplier")
@login_required
def search_supplier():

    q = request.args.get("q","")

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cur.execute("""
    SELECT id,name
    FROM suppliers
    WHERE name LIKE %s
    LIMIT 10
    """,('%'+q+'%',))

    data = cur.fetchall()

    cur.close()

    return jsonify(data)

@app.route("/api/add-supplier-quick", methods=["POST"])
@login_required
def add_supplier_quick():
    data = request.get_json()
    name = data.get("name")

    cur = mysql.connection.cursor()

    cur.execute("INSERT INTO suppliers (name, contact_person, created_at) VALUES (%s,%s,NOW())",
                (name,""))

    new_id = cur.lastrowid
    mysql.connection.commit()
    cur.close()

    return jsonify({"id": new_id, "name": name})

@app.route('/update-rate-logic', methods=['POST'])
@login_required
@owner_required
def update_rate_logic():
    product_id = request.form.get('product_id')
    new_price = request.form.get('new_price')

    if product_id and new_price:
        try:
            cur = mysql.connection.cursor()
            # 'products' table mein price update query
            sql = "UPDATE products SET selling_price = %s WHERE id = %s"
            cur.execute(sql, (new_price, product_id))

            mysql.connection.commit()
            cur.close()


        except Exception as e:
            flash(f"Error: {str(e)}", "danger")
    else:
        flash("Please provide all details!", "warning")

    return redirect(url_for('rate_update_page'))

@app.route("/estimate")
@login_required
def estimate():

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # =========================
    # PRODUCTS
    # =========================
    cur.execute("""
        SELECT id, product_name, selling_price, purchase_price
        FROM inventory
        ORDER BY product_name ASC
    """)
    rows = cur.fetchall()

    products = []
    for r in rows:
        products.append({
            "id": int(r["id"]) if r["id"] else 0,
            "product_name": r["product_name"] or "",
            "selling_price": float(r["selling_price"] or 0),
            "purchase_price": float(r["purchase_price"] or 0)
        })


    # =========================
    # HAMALS
    # =========================
    cur.execute("SELECT id, name FROM hamals ORDER BY name")
    hrows = cur.fetchall()

    hamals = []
    for h in hrows:
        hamals.append({
            "id": int(h["id"]),
            "name": h["name"]
        })


    # =========================
    # QUOTATION HISTORY
    # =========================
    cur.execute("""
        SELECT 
            e.id,
            e.customer_name,
            e.customer_mobile,
            e.total_amount,
            e.created_at,
            u.name AS staff
        FROM estimates e
        LEFT JOIN users u ON e.created_by = u.id
        ORDER BY e.created_at DESC
        LIMIT 50
    """)

    estimate_history = cur.fetchall()


    cur.close()

    return render_template(
        "estimate.html",
        products=products,
        hamals=hamals,
        estimate_history=estimate_history,   # ✅ HISTORY ADDED
        role=session.get("role"),
        current_date=date.today().strftime("%d %b %Y")
    )

@app.route("/estimate-pdf/<int:est_id>")
@login_required
def estimate_pdf(est_id):

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cur.execute("SELECT * FROM estimates WHERE id=%s", (est_id,))
    est = cur.fetchone()

    if not est:
        cur.close()
        return "Estimate not found", 404

    cur.execute("SELECT * FROM estimate_items WHERE estimate_id=%s",(est_id,))
    items = cur.fetchall()

    cur.close()

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)

    width, height = A4
    y = height - 40

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawCentredString(width/2, y, "ESTIMATE / QUOTATION")
    y -= 25

    pdf.setFont("Helvetica", 10)
    pdf.drawString(40, y, f"Customer: {est['customer_name']}")
    pdf.drawRightString(
        width - 40,
        y,
        f"Date: {est['created_at'].strftime('%d-%m-%Y')}"
    )
    y -= 25

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(40, y, "Item")
    pdf.drawRightString(380, y, "Qty")
    pdf.drawRightString(520, y, "Amount")
    y -= 10
    pdf.line(40, y, 550, y)
    y -= 15

    pdf.setFont("Helvetica", 10)

    for it in items:
        pdf.drawString(40, y, it["product_name"][:30])
        pdf.drawRightString(380, y, str(it["quantity"]))
        pdf.drawRightString(520, y, f"{it['amount']:.2f}")
        y -= 15

    y -= 10
    pdf.line(300, y, 550, y)
    y -= 20

    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawRightString(520, y, f"Total: ₹ {est['total_amount']:.2f}")

    pdf.showPage()
    pdf.save()

    buffer.seek(0)

    return Response(
        buffer,
        mimetype="application/pdf",
        headers={"Content-Disposition": f"inline; filename=estimate_{est_id}.pdf"}
    )

@app.route("/pending-bills")
@login_required
def pending_bills():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # ONLY DRAFT / PENDING BILLS
    cur.execute("""
        SELECT 
            id, 
            customer_name, 
            customer_mobile, 
            total_amount, 
            paid_amount, 
            balance_amount, 
            DATE_FORMAT(created_at, '%d %b %Y') as bill_date,
            DATEDIFF(CURDATE(), created_at) as days_pending,
            status
        FROM bills 
        WHERE status = 'PENDING'
        ORDER BY created_at DESC
    """)
    bills = cur.fetchall()

    # total draft amount
    cur.execute("""
        SELECT IFNULL(SUM(balance_amount),0) as total
        FROM bills
        WHERE status = 'PENDING'
    """)
    grand_total_pending = cur.fetchone()['total']

    cur.close()

    return render_template(
        "pending_bills.html",
        bills=bills,
        grand_total=grand_total_pending
    )

@app.route("/low-stock")
@login_required
def low_stock():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Query to find items where current stock is less than or equal to minimum level
    cur.execute("""
        SELECT 
            id, 
            product_name, 
            unit, 
            stock_quantity, 
            min_stock_level,
            purchase_price,
            (min_stock_level - stock_quantity) as requirement
        FROM inventory 
        WHERE stock_quantity <= min_stock_level
        ORDER BY stock_quantity ASC
    """)
    low_stock_items = cur.fetchall()

    # Count for the badge
    total_low_items = len(low_stock_items)

    cur.close()
    return render_template("low_stock.html", items=low_stock_items, total=total_low_items)


@app.route("/staff-salary")
@login_required
@owner_required
def staff_salary():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # 1. Fetch staff list for the management cards
    cur.execute("""
        SELECT id, name, username 
        FROM users 
        WHERE role != 'owner' 
        ORDER BY name
    """)
    staff_members = cur.fetchall()

    # 2. Fetch total pending advances per user
    cur.execute("""
        SELECT user_id, IFNULL(SUM(amount), 0) AS total_advance 
        FROM staff_advances 
        WHERE status = 'Pending' 
        GROUP BY user_id
    """)
    adv_rows = cur.fetchall()
    advances = {r["user_id"]: float(r["total_advance"]) for r in adv_rows}

    # 3. Fetch full payroll history for the table
    cur.execute("""
        SELECT 
            sp.id, 
            sp.base_salary, 
            sp.advance_deducted, 
            sp.net_paid, 
            sp.payout_date, 
            u.name 
        FROM staff_payouts sp 
        JOIN users u ON sp.user_id = u.id 
        ORDER BY sp.payout_date DESC
    """)
    history = cur.fetchall()

    cur.close()
    return render_template(
        "staff_salary.html",
        staff=staff_members,
        advances=advances,
        history=history
    )

@app.route("/generate-salary", methods=["POST"])
@login_required
@owner_required
def generate_salary():
    user_id = request.form.get("user_id")
    base_salary = float(request.form.get("base_salary") or 0)

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # 1. Get current pending advances (Total ₹1500 for Swaraj as per your image)
    cur.execute("""
        SELECT IFNULL(SUM(amount), 0) AS adv
        FROM staff_advances
        WHERE user_id = %s AND status = 'Pending'
    """, (user_id,))
    adv = float(cur.fetchone()["adv"])

    net_paid = base_salary - adv

    try:
        # 2. MATCHING YOUR DATABASE COLUMNS EXACTLY
        # We include 'net_paid' and 'advance_deducted' as seen in your MySQL Workbench
        cur.execute("""
            INSERT INTO staff_payouts 
            (user_id, amount, base_salary, advance_deducted, net_paid, payout_date, deductions) 
            VALUES (%s, %s, %s, %s, %s, NOW(), %s)
        """, (user_id, net_paid, base_salary, adv, net_paid, adv))

        payout_id = cur.lastrowid

        # 3. SETTLE ADVANCES (This is why it shows 0 on the next screen)
        cur.execute("""
            UPDATE staff_advances 
            SET status = 'Settled' 
            WHERE user_id = %s AND status = 'Pending'
        """, (user_id,))

        mysql.connection.commit()

        # Flash a success message

        return redirect(url_for("staff_salary"))

    except Exception as e:
        mysql.connection.rollback()
        # This will tell you if a column name is wrong
        print(f"Error saving to DB: {e}")

        return redirect(url_for("staff_salary"))
    finally:
        cur.close()


@app.route("/give-advance", methods=["POST"])
@login_required
@owner_required
def give_advance():

    user_id = request.form.get("user_id")
    amount = float(request.form.get("amount") or 0)
    reason = request.form.get("reason") or "Advance"

    cur = mysql.connection.cursor()

    cur.execute("""
        INSERT INTO staff_advances
        (user_id, amount, reason, status, created_at)
        VALUES (%s,%s,%s,'Pending',NOW())
    """, (user_id, amount, reason))

    mysql.connection.commit()
    cur.close()


    return redirect(url_for("staff_salary"))


@app.route("/print-salary-slip/<int:payout_id>")
@login_required
def print_salary_slip(payout_id):
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Payout details fetch karna staff ki details ke saath
    cur.execute("""
        SELECT sp.*, u.name, u.role, u.username 
        FROM staff_payouts sp 
        JOIN users u ON sp.user_id = u.id 
        WHERE sp.id = %s
    """, [payout_id])
    slip = cur.fetchone()
    cur.close()

    if not slip:

        return redirect(url_for('staff_salary'))

    return render_template("salary_slip_print.html", slip=slip)

@app.route("/api/search-hamal")
@login_required
def search_hamal():
    q = request.args.get("q","").strip()
    # DictCursor use karein taake 'name' key mile
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute("SELECT id, name FROM hamals WHERE name LIKE %s LIMIT 10", (f"%{q}%",))
    data = cur.fetchall()
    cur.close()
    return jsonify(data)


@app.route("/reports/daily-tally")
@login_required
def daily_tally():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    today = date.today()

    # 1. Sale Analysis
    cur.execute("""
        SELECT payment_mode, SUM(total_amount) as total 
FROM bills WHERE DATE(created_at)=%s AND status!='ESTIMATE'
        GROUP BY payment_mode
    """, (today,))
    sales = cur.fetchall()

    # 2. Expenses (Jawak)
    cur.execute("SELECT IFNULL(SUM(amount),0) as total FROM expenses WHERE DATE(created_at) = %s", (today,))
    expenses = cur.fetchone()['total']

    # 3. Galla Calculation (Cash in Hand)
    cur.execute("""
        SELECT (
            (SELECT IFNULL(SUM(paid_amount),0) FROM bills WHERE payment_mode='Cash' AND DATE(created_at)=%s) - 
            (SELECT IFNULL(SUM(amount),0) FROM expenses WHERE DATE(created_at)=%s)
        ) as cash_in_hand
    """, (today, today))
    galla = cur.fetchone()['cash_in_hand']

    # ================= PROFIT TODAY =================
    cur.execute("""
        SELECT IFNULL(SUM((bi.rate - bi.purchase_price) * bi.quantity),0) AS profit
        FROM bill_items bi
        JOIN bills b ON bi.bill_id = b.id
        WHERE DATE(b.created_at)=%s
        AND b.status IN ('DONE','BAKI')
    """,(today,))

    today_profit = cur.fetchone()['profit']
    cur.close()
    return render_template("daily_tally.html",
                           sales=sales, expenses=expenses, galla=galla,profit=today_profit)

@app.route('/hamal-account', methods=['GET', 'POST'])
@login_required
def hamal_account():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    if request.method == 'POST':
        name = request.form.get('name')
        mobile = request.form.get('mobile')

        try:
            # Using 'total_due' as seen in your MySQL screenshot
            cur.execute("""
                INSERT INTO hamals (name, mobile, total_due) 
                VALUES (%s, %s, %s)
            """, (name, mobile, 0.00))

            mysql.connection.commit()

        except Exception as e:
            mysql.connection.rollback()


        # This redirect prevents the "multiples" issue when going back
        return redirect(url_for('hamal_account'))

    # GET logic: Fetch using actual column names
    cur.execute("SELECT id, name, mobile, total_due FROM hamals ORDER BY id DESC")
    hamal_list = cur.fetchall()
    cur.close()

    return render_template('hamal_account.html', hamals=hamal_list)


# Edit Route
@app.route("/edit-hamal/<int:id>", methods=["POST"])
@login_required
def edit_hamal(id):
    name = request.form.get("name")
    mobile = request.form.get("mobile")
    cur = mysql.connection.cursor()
    cur.execute("UPDATE hamals SET name=%s, mobile=%s WHERE id=%s", (name, mobile, id))
    mysql.connection.commit()
    cur.close()

    return redirect(url_for('hamal_account'))

# Delete Route
@app.route("/delete-hamal/<int:id>")
@login_required
def delete_hamal(id):
    cur = mysql.connection.cursor()
    # FIXED: hamal_transactions ki jagah hamal_accounts use karein
    cur.execute("DELETE FROM hamal_ledger WHERE hamal_id=%s", (id,))
    cur.execute("DELETE FROM hamals WHERE id=%s", (id,))
    mysql.connection.commit()
    cur.close()

    return redirect(url_for('hamal_account'))


@app.route('/hamal-ledger/<int:hamal_id>')
@login_required
def hamal_ledger(hamal_id):

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cur.execute("SELECT * FROM hamals WHERE id=%s",(hamal_id,))
    hamal = cur.fetchone()

    cur.execute("""
        SELECT *
        FROM hamal_ledger
        WHERE hamal_id=%s
        ORDER BY created_at DESC
    """,(hamal_id,))

    rows = cur.fetchall()
    cur.close()

    ledger = []
    bill_map = {}

    for r in rows:

        # CASH PAYMENT
        if r["debit_amount"] > 0:

            ledger.append({
                "created_at": r["created_at"],
                "bill_id": None,
                "description": "Cash Payment",
                "hamali": 0,
                "bhada": 0,
                "credit": 0,
                "debit": r["debit_amount"],
                "balance": r["balance_amount"]
            })

            continue


        bill_id = r["bill_id"]

        if bill_id not in bill_map:

            bill_map[bill_id] = {
                "created_at": r["created_at"],
                "bill_id": bill_id,
                "description": "",
                "hamali": 0,
                "bhada": 0,
                "credit": 0,
                "debit": 0,
                "balance": r["balance_amount"]
            }

            ledger.append(bill_map[bill_id])


        if "hamali" in r["description"].lower():
            bill_map[bill_id]["hamali"] += r["credit_amount"]

        if "bhada" in r["description"].lower():
            bill_map[bill_id]["bhada"] += r["credit_amount"]

        bill_map[bill_id]["credit"] += r["credit_amount"]
        bill_map[bill_id]["balance"] = r["balance_amount"]

    return render_template(
        "hamal_ledger.html",
        hamal=hamal,
        transactions=ledger
    )


@app.route('/pay-hamal/<int:hamal_id>', methods=['POST'])
@login_required
def pay_hamal(hamal_id):
    amount = float(request.form.get('amount'))
    description = request.form.get('description') or "Cash Payment"

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    try:
        # 1. Hamal ka current balance fetch karna
        cur.execute("SELECT total_due FROM hamals WHERE id = %s", (hamal_id,))
        current_due = cur.fetchone()['total_due']

        # New balance calculation (Payment se due kam hota hai)
        new_balance = float(current_due) - amount

        # 2. hamal_ledger mein entry insert karna
        # debit_amount = payment amount, credit_amount = 0
        cur.execute("""
            INSERT INTO hamal_ledger (hamal_id, description, debit_amount, credit_amount, balance_amount, created_at)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (hamal_id, description, amount, 0, new_balance, datetime.datetime.now()))

        # 3. main hamals table mein balance update karna
        cur.execute("UPDATE hamals SET total_due = %s WHERE id = %s", (new_balance, hamal_id))

        mysql.connection.commit()

    except Exception as e:
        mysql.connection.rollback()

    finally:
        cur.close()

    return redirect(url_for('hamal_account'))

@app.route("/upi-balance")
@login_required
def upi_balance_report():

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    filter_type = request.args.get("type", "today")

    date_condition = ""

    if filter_type == "today":
        date_condition = "DATE(created_at)=CURDATE()"

    elif filter_type == "week":
        date_condition = "YEARWEEK(created_at,1)=YEARWEEK(CURDATE(),1)"

    elif filter_type == "month":
        date_condition = "MONTH(created_at)=MONTH(CURDATE()) AND YEAR(created_at)=YEAR(CURDATE())"

    else:
        date_condition = "1=1"

    # UPI RECEIVED FROM BILLS
    cur.execute(f"""
        SELECT IFNULL(SUM(paid_amount),0) AS total
        FROM bills
        WHERE upi_account IN ('Owner','Staff')
        AND {date_condition}
    """)
    bill_received = cur.fetchone()['total']

    # UPI RECEIVED FROM ADVANCE BOOKINGS
    cur.execute(f"""
        SELECT IFNULL(SUM(advance_amount),0) AS total
        FROM advance_bookings
        WHERE payment_mode IN ('Owner_UPI','Staff_UPI')
        AND {date_condition}
    """)
    advance_received = cur.fetchone()['total']
    # UPI RECEIVED FROM CHILLAR
    cur.execute(f"""
        SELECT IFNULL(SUM(amount),0) AS total
        FROM chillar_entries
        WHERE payment_mode IN ('Owner','Staff','UPI')
        AND type='BIKRI'
        AND {date_condition}
    """)
    chillar_received = cur.fetchone()['total']

    # TOTAL UPI RECEIVED
    received = float(bill_received) + float(advance_received) + float(chillar_received)    # UPI SPENT
    cur.execute(f"""
        SELECT IFNULL(SUM(amount),0) AS total
        FROM supplier_payments
        WHERE payment_method='UPI'
        AND DATE(payment_date)=CURDATE()
    """)
    spent = cur.fetchone()['total']

    balance = float(received) - float(spent)

    cur.close()

    return render_template(
        "upi_balance.html",
        received=received,
        spent=spent,
        balance=balance,
        filter_type=filter_type
    )

@app.route("/add-expense", methods=["POST"])
@login_required
def add_expense():

    amount = float(request.form.get("amount") or 0)
    description = request.form.get("description")
    date_val = request.form.get("date")

    cur = mysql.connection.cursor()

    cur.execute("""
        INSERT INTO expenses
        (amount, description, created_at)
        VALUES (%s,%s,%s)
    """,(amount,description,date_val))

    mysql.connection.commit()
    cur.close()

    return redirect(url_for("daily_report"))

@app.route("/upi-report")
@login_required
def upi_report():

    filter_type = request.args.get("type", "day")
    page = int(request.args.get("page", 1))
    per_page = 50
    offset = (page - 1) * per_page
    search = request.args.get("search", "").strip()

    today = date.today()
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # ================= DATE FILTER =================
    params = []

    if filter_type == "day":
        date_filter_b = "DATE(b.created_at) = %s"
        date_filter_c = "DATE(c.created_at) = %s"
        params.append(today)

    elif filter_type == "month":
        date_filter_b = "YEAR(b.created_at)=YEAR(CURDATE()) AND MONTH(b.created_at)=MONTH(CURDATE())"
        date_filter_c = "YEAR(c.created_at)=YEAR(CURDATE()) AND MONTH(c.created_at)=MONTH(CURDATE())"

    elif filter_type == "year":
        date_filter_b = "YEAR(b.created_at)=YEAR(CURDATE())"
        date_filter_c = "YEAR(c.created_at)=YEAR(CURDATE())"

    elif filter_type == "week":
        date_filter_b = "YEARWEEK(b.created_at,1)=YEARWEEK(CURDATE(),1)"
        date_filter_c = "YEARWEEK(c.created_at,1)=YEARWEEK(CURDATE(),1)"

    else:
        date_filter_b = "1=1"
        date_filter_c = "1=1"

    # ================= SEARCH =================
    search_filter = ""
    search_params = []

    if search:
        search_filter = "AND (customer_name LIKE %s OR CAST(bill_no AS CHAR) LIKE %s)"
        search_params = [f"%{search}%", f"%{search}%"]

    # ================= BASE UNION =================
    base_query = f"""
SELECT 
    b.id AS bill_no,
    b.customer_name,
    (b.paid_amount - IFNULL(ab.advance_amount,0)) AS total_amount,
    b.upi_account,
    u.name AS staff_name,
    b.created_at,
    'bill' AS source
FROM bills b
LEFT JOIN users u ON b.upi_staff_id = u.id
LEFT JOIN advance_bookings ab ON b.advance_id = ab.id
WHERE b.upi_account IN ('Owner','Staff')
  AND b.paid_amount > 0
  AND b.status IN ('DONE','BAKI')
  AND {date_filter_b}
        UNION ALL

        SELECT
            c.id AS bill_no,
            c.customer_name,
            c.amount AS total_amount,
            c.payment_mode AS upi_account,
            NULL AS staff_name,
            c.created_at,
            'chillar' AS source
        FROM chillar_entries c
        WHERE c.payment_mode IN ('Owner','Staff')
        AND c.type='BIKRI'
        AND {date_filter_c}
                UNION ALL

       SELECT
        ab.id AS bill_no,
        cu.name AS customer_name,
        ab.advance_amount AS total_amount,
        ab.payment_mode AS upi_account,
        NULL AS staff_name,
        ab.created_at,
        'advance' AS source
        FROM advance_bookings ab
        LEFT JOIN customers cu ON ab.customer_id = cu.id
        WHERE ab.payment_mode IN ('Owner_UPI','Staff_UPI')
    """
    # IMPORTANT: duplicate params for UNION (bills + chillar)
    base_params = params + params

    # ================= COUNT =================
    count_query = f"""
        SELECT COUNT(*) AS total FROM (
            {base_query}
        ) AS combined
        WHERE 1=1 {search_filter}
    """

    cur.execute(count_query, base_params + search_params)
    total_rows = cur.fetchone()["total"]
    total_pages = (total_rows // per_page) + (1 if total_rows % per_page else 0)

    # ================= DATA =================
    data_query = f"""
        SELECT * FROM (
            {base_query}
        ) AS combined
        WHERE 1=1 {search_filter}
        ORDER BY created_at DESC
        LIMIT %s OFFSET %s
    """

    cur.execute(data_query, base_params + search_params + [per_page, offset])
    transactions = cur.fetchall()

    # ================= TOTAL =================
    total_query = f"""
        SELECT IFNULL(SUM(total_amount),0) AS total FROM (
            {base_query}
        ) AS combined
        WHERE 1=1 {search_filter}
    """

    cur.execute(total_query, base_params + search_params)
    total_upi = cur.fetchone()["total"]

    cur.close()

    return render_template(
        "upi_report.html",
        transactions=transactions,
        total_upi=total_upi,
        page=page,
        total_pages=total_pages,
        filter_type=filter_type,
        search=search
    )

@app.route("/cash-tally")
@login_required
def cash_tally():

    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")

    if not from_date:
        from_date = date.today().strftime("%Y-%m-%d")

    if not to_date:
        to_date = date.today().strftime("%Y-%m-%d")

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    transaction_list = []

    # =========================
    # CASH FROM BILLS
    # =========================
    cur.execute("""
    SELECT 
        DATE(b.created_at) as date,
        CONCAT('Bill #',b.id) as description,
        'INCOME' as type,
        CASE 
            WHEN b.advance_id IS NOT NULL THEN 0
            ELSE IFNULL(b.paid_amount,0)
        END as amount
    FROM bills b
    WHERE DATE(b.created_at) BETWEEN %s AND %s
    AND b.upi_account='Cash'
    AND b.status IN ('DONE','BAKI')
    """, (from_date, to_date))
    transaction_list += cur.fetchall()


    # =========================
    # CHILLAR CASH
    # =========================
    cur.execute("""
    SELECT 
        DATE(created_at) as date,
        'Chillar Entry' as description,
        'INCOME' as type,
        amount
    FROM chillar_entries
    WHERE DATE(created_at) BETWEEN %s AND %s
    AND payment_mode='Cash'
    """,(from_date,to_date))

    transaction_list += cur.fetchall()


    # =========================
    # ADVANCE BOOKINGS CASH
    # =========================
    cur.execute("""
    SELECT 
        DATE(created_at) as date,
        'Advance Booking' as description,
        'INCOME' as type,
        advance_amount as amount
    FROM advance_bookings
    WHERE DATE(created_at) BETWEEN %s AND %s
    AND payment_mode='Cash'
    """,(from_date,to_date))

    transaction_list += cur.fetchall()


    # =========================
    # STAFF SALARY (CASH)
    # =========================
    cur.execute("""
    SELECT 
        DATE(payout_date) as date,
        'Staff Salary' as description,
        'EXPENSE' as type,
        amount
    FROM staff_payouts
    WHERE DATE(payout_date) BETWEEN %s AND %s
    """,(from_date,to_date))

    transaction_list += cur.fetchall()


    # =========================
    # EXPENSES
    # =========================
    cur.execute("""
    SELECT 
        DATE(created_at) as date,
        'Expense' as description,
        'EXPENSE' as type,
        amount
    FROM expenses
    WHERE DATE(created_at) BETWEEN %s AND %s
    """,(from_date,to_date))

    transaction_list += cur.fetchall()

    # =========================
    # SUPPLIER PAYMENTS (CASH)
    # =========================
    cur.execute("""
    SELECT 
        DATE(payment_date) as date,
        'Supplier Payment' as description,
        'EXPENSE' as type,
        amount
    FROM supplier_payments
    WHERE payment_method='Cash'
    AND DATE(payment_date) BETWEEN %s AND %s
    """, (from_date, to_date))

    transaction_list += cur.fetchall()

    # =========================
    # HAMAL PAYMENTS
    # =========================
    cur.execute("""
    SELECT 
        DATE(created_at) as date,
        'Hamal Payment' as description,
        'EXPENSE' as type,
        debit_amount as amount
    FROM hamal_ledger
    WHERE debit_amount > 0
    AND DATE(created_at) BETWEEN %s AND %s
    """, (from_date, to_date))

    transaction_list += cur.fetchall()

    # =========================
    # CALCULATE CASH
    # =========================
    cash_today = sum(float(x['amount']) for x in transaction_list if x['type'] == 'INCOME')
    today_expenses = sum(float(x['amount']) for x in transaction_list if x['type'] == 'EXPENSE')

    cash_collection = cash_today - today_expenses


    # =========================
    # ACTUAL CASH CHECK
    # =========================
    actual_cash = request.args.get("actual_cash")

    cash_difference = None

    if actual_cash:
        actual_cash = float(actual_cash)
        cash_difference = actual_cash - cash_collection


    # =========================
    # SORT TRANSACTIONS
    # =========================
    transaction_list.sort(key=lambda x: (x['date'], x['type']), reverse=True)

    cur.close()

    return render_template(
        "cash_tally.html",
        transactions=transaction_list,
        cash_today=cash_today,
        today_expenses=today_expenses,
        cash_collection=cash_collection,
        cash_difference=cash_difference,
        from_date=from_date,
        to_date=to_date
    )

@app.route("/close-day", methods=["POST"])
@login_required
@owner_required
def close_day():
    print("CLOSE DAY CLICKED")  # ADD THIS
    cur = mysql.connection.cursor()

    # check if already closed
    cur.execute("""
        SELECT id FROM daily_closing
        WHERE closing_date = CURDATE()
    """)

    if cur.fetchone():

        return redirect(url_for("closing_summary"))
    # =========================
    # TOTAL SALE
    # =========================
    cur.execute("""
        SELECT IFNULL(SUM(total_amount),0)
        FROM bills
        WHERE DATE(created_at)=CURDATE()
        AND status IN ('DONE','BAKI')
    """)
    bill_sale = cur.fetchone()[0] or 0

    cur.execute("""
        SELECT IFNULL(SUM(amount),0)
        FROM chillar_entries
        WHERE DATE(created_at)=CURDATE()
        AND type='BIKRI'
    """)
    chillar_sale = cur.fetchone()[0] or 0

    total_sale = float(bill_sale) + float(chillar_sale)

    # =========================
    # CASH COLLECTION
    # =========================
    cur.execute("""
        SELECT IFNULL(SUM(paid_amount),0)
        FROM bills
        WHERE DATE(created_at)=CURDATE()
        AND upi_account='Cash'
        AND status IN ('DONE','BAKI')
    """)
    cash_bills = cur.fetchone()[0] or 0

    cur.execute("""
        SELECT IFNULL(SUM(amount),0)
        FROM chillar_entries
        WHERE DATE(created_at)=CURDATE()
        AND payment_mode='Cash'
    """)
    chillar_cash = cur.fetchone()[0] or 0

    cash_collection = float(cash_bills) + float(chillar_cash)

    # =========================
    # UPI COLLECTION
    # =========================
    cur.execute("""
        SELECT IFNULL(SUM(paid_amount),0)
        FROM bills
        WHERE DATE(created_at)=CURDATE()
        AND upi_account IN ('Owner','Staff')
        AND status IN ('DONE','BAKI')
    """)
    upi_bills = cur.fetchone()[0] or 0

    cur.execute("""
        SELECT IFNULL(SUM(amount),0)
        FROM chillar_entries
        WHERE DATE(created_at)=CURDATE()
        AND payment_mode='UPI'
    """)
    upi_chillar = cur.fetchone()[0] or 0

    upi_collection = float(upi_bills) + float(upi_chillar)

    total_collection = cash_collection + upi_collection

    # =========================
    # EXPENSES
    # =========================
    cur.execute("""
        SELECT IFNULL(SUM(amount),0)
        FROM expenses
        WHERE DATE(created_at)=CURDATE()
    """)

    total_expenses = cur.fetchone()[0] or 0

    # =========================
    # MARKET PENDING
    # =========================
    cur.execute("""
        SELECT IFNULL(SUM(balance_amount),0)
        FROM bills
        WHERE balance_amount>0
        AND status IN ('DONE','BAKI')
    """)

    market_pending = cur.fetchone()[0] or 0

    # =========================
    # PROFIT
    # =========================
    cur.execute("""
        SELECT IFNULL(SUM((bi.rate - bi.purchase_price)*bi.quantity),0)
        FROM bill_items bi
        JOIN bills b ON bi.bill_id=b.id
        WHERE DATE(b.created_at)=CURDATE()
        AND b.status IN ('DONE','BAKI')
    """)

    total_profit = cur.fetchone()[0] or 0

    # =========================
    # CASH IN GALLA
    # =========================
    cash_in_galla = float(cash_collection) - float(total_expenses)

    # =========================
    # SAVE SNAPSHOT
    # =========================
    cur.execute("""
        INSERT INTO daily_closing
        (
            closing_date,
            total_sale,
            cash_collection,
            upi_collection,
            total_collection,
            total_expenses,
            cash_in_galla,
            total_profit,
            market_pending,
            closed_by
        )
        VALUES
        (CURDATE(),%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        total_sale,
        cash_collection,
        upi_collection,
        total_collection,
        total_expenses,
        cash_in_galla,
        total_profit,
        market_pending,
        session["user_id"]
    ))

    mysql.connection.commit()
    cur.close()


    return redirect(url_for("closing_summary"))

@app.route("/closing-summary")
@login_required
def closing_summary():

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cur.execute("""
        SELECT dc.*, u.name AS closed_by_name
        FROM daily_closing dc
        LEFT JOIN users u ON dc.closed_by = u.id
        ORDER BY dc.closing_date DESC
        LIMIT 1
    """)

    data = cur.fetchone()   # fetch record

    cur.close()

    if data is None:
        flash("No closing data found.")
        return redirect(url_for("dashboard"))

    return render_template(
        "closing_summary.html",
        data=data
    )

# Yeh line left side se shuru honi chahiye (No indentation)
@app.route("/logout")
def logout():
    session.clear()

    return redirect(url_for("login"))


if __name__ == "__main__":
    app.run(debug=True)